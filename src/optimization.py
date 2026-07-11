"""RouteTech Faz 2 ("Gelişmiş Çözüm Aşaması") CP-SAT optimizasyon modeli.

main.py bu betiği subprocess olarak çalıştırır; girdi olarak
ROUTETECH_OPTIMIZATION_INPUT ortam değişkeniyle verilen optimization_input.json'u
okur (src/submission.py tarafından üretilir: forecast, route_matrix, vehicle_costs,
rental_limits, handling_capacity, tir_capacity).

Karar değişkeni tasarımı, Faz-1'de kanıtlanmış aggregate (hat×gün×araç_türü, tek
tek araç kimliği yok) yaklaşımın saat/dispatch_slot boyutuyla genişletilmiş hali:
    kiralik_x[(hat, gun, slot, arac_turu)]  — o slotta kalkan kiralık araç sayısı
    spot_y[(hat, gun, slot, arac_turu)]     — o slotta kalkan spot araç sayısı
    ertelenen_talep[(hat, gun, slot)]       — bir sonraki slota ertelenen talep

Zaman ekseni serbest bir CP-SAT değişkeni DEĞİL — kalkış (gun, slot) sabit bir
epok kümesinden seçiliyor, varış/elleçleme zamanları src/time_model.py ile
analitik olarak (girdi verisinden) türetiliyor. Bu, durum uzayının patlamasını
önleyen kilit tasarım kararı (bkz. plan dosyası).

Kapsam notu (Stage B): serbest konsolidasyon (PDF: "Her transfer merkezinde
kısıtlamalara uygun konsolidasyon yapabilirsiniz") bu sürümde MODELLENMİYOR —
yalnızca direkt hat sevkiyatı var. Konsolidasyon/aktarma esnekliği Stage C'de
ALNS repair operatörleriyle eklenecek.
"""

from __future__ import annotations

import multiprocessing
import os
import sys
import json
from pathlib import Path

import pandas as pd
from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.alns.cost_model import vehicle_leg_cost  # noqa: E402
from src.alns.time_model import (  # noqa: E402
    DISPATCH_SLOTS,
    arrival_day,
    build_route_lookup,
    sla_cezasi_tl,
    slot_to_hour,
)

# ============================================================================
# 1. ORTAM DEĞİŞKENLERİ
# ============================================================================
ENV_MAX_TIME = float(os.environ.get("ROUTETECH_MAX_TIME_SECONDS", "300"))
ENV_LOG_PROGRESS = os.environ.get("ROUTETECH_LOG_SEARCH_PROGRESS", "1") == "1"

DEFAULT_INPUT_JSON = PROJECT_ROOT / "data" / "outputs" / "optimization_input.json"
INPUT_JSON_PATH = Path(os.environ.get("ROUTETECH_OPTIMIZATION_INPUT", str(DEFAULT_INPUT_JSON)))

MAX_SPOT = 500

# ============================================================================
# 2. GİRDİYİ OKU
# ============================================================================
if not INPUT_JSON_PATH.exists():
    raise FileNotFoundError(
        f"Optimizasyon girdisi bulunamadı: {INPUT_JSON_PATH}. Once `python main.py` calistirilmali."
    )

with INPUT_JSON_PATH.open("r", encoding="utf-8") as f:
    payload = json.load(f)

df_forecast = pd.DataFrame(payload["forecast"])
df_route_matrix = pd.DataFrame(payload["route_matrix"])
df_vehicle_costs = pd.DataFrame(payload["vehicle_costs"])
df_rental_limits = pd.DataFrame(payload["rental_limits"])
df_handling_capacity = pd.DataFrame(payload["handling_capacity"])
df_tir_capacity = pd.DataFrame(payload["tir_capacity"])

if df_forecast.empty:
    raise ValueError(f"Forecast payload bos: {INPUT_JSON_PATH}")

route_lookup = build_route_lookup(df_route_matrix)

arac_turleri = df_vehicle_costs["vehicle_type"].tolist()
tir_arac_turu = next(
    (v for v in arac_turleri if str(v).strip().casefold() in ("tır", "tir")), None
)

arac_parametreleri = {}
for row in df_vehicle_costs.itertuples():
    arac_parametreleri[row.vehicle_type] = {
        "kapasite_desi": float(row.capacity),
        "rental_hourly": float(row.rental_hourly),
        "rental_km": float(row.rental_km),
        "spot_hourly": float(row.spot_hourly),
        "spot_km": float(row.spot_km),
    }

kiralik_stok_gunluk: dict[tuple, int] = {}
for row in df_rental_limits.itertuples():
    hat = (row.source, row.destination)
    key = (hat, row.vehicle_type)
    kiralik_stok_gunluk[key] = kiralik_stok_gunluk.get(key, 0) + int(row.vehicle_count)

handling_capacity = {row.center: float(row.capacity) for row in df_handling_capacity.itertuples()}
tir_capacity = {row.center: float(row.capacity) for row in df_tir_capacity.itertuples()}

# ============================================================================
# 3. TALEP / HAT / GÜN / SLOT ÇIKARIMI
# ============================================================================
df_forecast = df_forecast.copy()
df_forecast["date"] = pd.to_datetime(df_forecast["date"])
df_forecast["gun_key"] = df_forecast["date"].dt.strftime("%Y-%m-%d")
df_forecast["slot"] = df_forecast["slot"].astype(str)

gunler = sorted(df_forecast["gun_key"].unique())
hatlar = sorted({(r.source, r.destination) for r in df_forecast.itertuples()})
merkezler = sorted(set(handling_capacity) | set(tir_capacity))

talep_verisi: dict[tuple, float] = {}
for row in df_forecast.itertuples():
    hat = (row.source, row.destination)
    key = (hat, row.gun_key, row.slot)
    talep_verisi[key] = talep_verisi.get(key, 0.0) + max(0.0, float(row.recommended_demand))

# Kronolojik (gun, slot) sıralaması — ertelenen talebin zincirlenmesi ve saatlik
# SLA cezası bu sıralamaya göre hesaplanır.
zaman_sirali: list[tuple[str, str]] = [(g, s) for g in gunler for s in DISPATCH_SLOTS]
zaman_index = {gs: i for i, gs in enumerate(zaman_sirali)}


def _adim_saat_farki(g: str, s: str) -> float:
    """(g, s)'den zaman_sirali'ndaki bir sonraki adıma kadar geçen saat (yoksa 0)."""
    idx = zaman_index[(g, s)]
    if idx + 1 >= len(zaman_sirali):
        return 0.0
    g2, s2 = zaman_sirali[idx + 1]
    gun_farki = (pd.Timestamp(g2) - pd.Timestamp(g)).days
    return float(gun_farki * 24 + (slot_to_hour(s2) - slot_to_hour(s)))


def _varis_gunu(hat: tuple, gun: str, slot: str, arac_turu: str) -> str | None:
    return arrival_day(route_lookup, gunler, hat, gun, slot, arac_turu)


def _maliyet_katsayisi(hat: tuple, arac_turu: str, hourly: float, km_maliyet: float) -> int:
    return vehicle_leg_cost(route_lookup, hat, arac_turu, hourly, km_maliyet)


# ============================================================================
# 4. MODEL VE KARAR DEĞİŞKENLERİ
# ============================================================================
model = cp_model.CpModel()

max_talep_hatta = max(1, int(round(sum(talep_verisi.values()))))

kiralik_x = {}
spot_y = {}
ertelenen_talep = {}
biriken_talep = {}
kiralik_yuk_by_type = {}
spot_yuk_by_type = {}
kiralik_tasinan_yuk_dict = {}

for h in hatlar:
    for (g, s) in zaman_sirali:
        for a in arac_turleri:
            stok = kiralik_stok_gunluk.get((h, a), 0)
            kiralik_x[(h, g, s, a)] = model.NewIntVar(0, stok, f"kiralik_{h}_{g}_{s}_{a}")
            spot_y[(h, g, s, a)] = model.NewIntVar(0, MAX_SPOT, f"spot_{h}_{g}_{s}_{a}")
        ertelenen_talep[(h, g, s)] = model.NewIntVar(0, max_talep_hatta, f"ert_{h}_{g}_{s}")
        biriken_talep[(h, g, s)] = model.NewIntVar(0, max_talep_hatta, f"bir_{h}_{g}_{s}")

# Kiralık stok günlük toplamdır — iki slota bölünse de günlük toplamı aşamaz.
for h in hatlar:
    for g in gunler:
        for a in arac_turleri:
            stok = kiralik_stok_gunluk.get((h, a), 0)
            model.Add(sum(kiralik_x[(h, g, s, a)] for s in DISPATCH_SLOTS) <= stok)

# ============================================================================
# 5. KISITLAR — Talep Dengesi / Yük Dağıtımı / %10 Doluluk
# ============================================================================
for h in hatlar:
    for (g, s) in zaman_sirali:
        idx = zaman_index[(g, s)]

        # --- Talep Dengesi (erteleme zinciri) ---
        bugun_talep = int(round(talep_verisi.get((h, g, s), 0.0)))
        if idx == 0:
            model.Add(biriken_talep[(h, g, s)] == bugun_talep)
        else:
            onceki_g, onceki_s = zaman_sirali[idx - 1]
            model.Add(
                biriken_talep[(h, g, s)]
                == ertelenen_talep[(h, onceki_g, onceki_s)] + bugun_talep
            )

        # --- Kiralık net yük (araç türü bazında, varış-günü muhasebesi için) ---
        kiralik_yuk_terimleri = []
        for a in arac_turleri:
            kap = int(arac_parametreleri[a]["kapasite_desi"])
            yuk_a = model.NewIntVar(0, kiralik_stok_gunluk.get((h, a), 0) * kap, f"kir_net_{h}_{g}_{s}_{a}")
            kiralik_yuk_by_type[(h, g, s, a)] = yuk_a
            kiralik_yuk_terimleri.append(yuk_a)
            model.Add(yuk_a <= kiralik_x[(h, g, s, a)] * kap)

        kiralik_tasinan_yuk = model.NewIntVar(0, max_talep_hatta, f"kir_toplam_{h}_{g}_{s}")
        model.Add(kiralik_tasinan_yuk == cp_model.LinearExpr.Sum(kiralik_yuk_terimleri))
        kiralik_tasinan_yuk_dict[(h, g, s)] = kiralik_tasinan_yuk

        # --- Spot net yük (araç türü bazında, %10 doluluk kuralı) ---
        spot_tasinan_yuk_listesi = []
        for a in arac_turleri:
            kap = int(arac_parametreleri[a]["kapasite_desi"])
            tasinan_yuk_a = model.NewIntVar(0, MAX_SPOT * kap, f"spot_net_{h}_{g}_{s}_{a}")
            spot_tasinan_yuk_listesi.append(tasinan_yuk_a)
            spot_yuk_by_type[(h, g, s, a)] = tasinan_yuk_a
            model.Add(tasinan_yuk_a <= spot_y[(h, g, s, a)] * kap)

            # Son (gün, slot) hariç: spot araç en az %10 dolu olmalı.
            if idx != len(zaman_sirali) - 1:
                model.Add(spot_y[(h, g, s, a)] * kap <= tasinan_yuk_a * 10)

        # --- Yük dağıtım dengesi ---
        model.Add(
            biriken_talep[(h, g, s)]
            == kiralik_tasinan_yuk + cp_model.LinearExpr.Sum(spot_tasinan_yuk_listesi)
            + ertelenen_talep[(h, g, s)]
        )
        model.Add(ertelenen_talep[(h, g, s)] <= biriken_talep[(h, g, s)])

# --- Son (gün, slot): erteleme yasağı ---
son_gun, son_slot = zaman_sirali[-1]
for h in hatlar:
    model.Add(ertelenen_talep[(h, son_gun, son_slot)] == 0)

# ============================================================================
# 6. KISITLAR — Elleçleme Kapasitesi ve Tır Kapasitesi (Faz 2, YENİ)
#     Elleçleme: desi bazlı, TM+gün. Çıkış elleçlemesi kalkış (g,s) gününe,
#     varış elleçlemesi ise ARAÇ TÜRÜNE göre değişen seyir süresiyle hesaplanan
#     gerçek varış gününe düşer (bkz. _varis_gunu — analitik, CP-SAT değişkeni değil).
# ============================================================================
for tm in merkezler:
    capacity = handling_capacity.get(tm)
    if capacity is None:
        continue
    for g in gunler:
        terimler = []

        for h in hatlar:
            src, dst = h
            if src == tm:
                for s in DISPATCH_SLOTS:
                    terimler.append(kiralik_tasinan_yuk_dict[(h, g, s)])
                    terimler.extend(spot_yuk_by_type[(h, g, s, a)] for a in arac_turleri)

            if dst == tm:
                for g_kalkis in gunler:
                    for s in DISPATCH_SLOTS:
                        for a in arac_turleri:
                            if _varis_gunu(h, g_kalkis, s, a) != g:
                                continue
                            terimler.append(kiralik_yuk_by_type[(h, g_kalkis, s, a)])
                            terimler.append(spot_yuk_by_type[(h, g_kalkis, s, a)])

        if terimler:
            model.Add(cp_model.LinearExpr.Sum(terimler) <= int(capacity))

# --- Tır kapasitesi (araç SAYISI bazlı, yalnızca Tır türü, giden+gelen toplam) ---
if tir_arac_turu is not None:
    for tm in merkezler:
        capacity = tir_capacity.get(tm)
        if capacity is None:
            continue
        for g in gunler:
            terimler = []

            for h in hatlar:
                src, dst = h
                if src == tm:
                    for s in DISPATCH_SLOTS:
                        terimler.append(kiralik_x[(h, g, s, tir_arac_turu)])
                        terimler.append(spot_y[(h, g, s, tir_arac_turu)])

                if dst == tm:
                    for g_kalkis in gunler:
                        for s in DISPATCH_SLOTS:
                            if _varis_gunu(h, g_kalkis, s, tir_arac_turu) != g:
                                continue
                            terimler.append(kiralik_x[(h, g_kalkis, s, tir_arac_turu)])
                            terimler.append(spot_y[(h, g_kalkis, s, tir_arac_turu)])

            if terimler:
                model.Add(cp_model.LinearExpr.Sum(terimler) <= int(capacity))

# ============================================================================
# 7. AMAÇ FONKSİYONU — Saatlik Araç Maliyeti + Saatlik SLA Cezası
# ============================================================================
maliyet_kalemleri = []
kiralik_sabit_toplam = 0

for h in hatlar:
    for a in arac_turleri:
        p = arac_parametreleri[a]
        stok = kiralik_stok_gunluk.get((h, a), 0)
        if stok <= 0:
            continue
        # Kiralık araçlar zorunlu kalkış — maliyet, kullanılan slottan bağımsız,
        # stoğa göre her gün sabit ödenir (Faz 1'deki tasarımla aynı mantık).
        birim_maliyet = _maliyet_katsayisi(h, a, p["rental_hourly"], p["rental_km"])
        kiralik_sabit_toplam += len(gunler) * stok * birim_maliyet

    for a in arac_turleri:
        p = arac_parametreleri[a]
        birim_maliyet = _maliyet_katsayisi(h, a, p["spot_hourly"], p["spot_km"])
        for (g, s) in zaman_sirali:
            maliyet_kalemleri.append(spot_y[(h, g, s, a)] * birim_maliyet)

    # BILINEN SINIRLAMA: bu ceza, ertelenen_talep'in HER adimda (talebin gercek
    # hedef_teslim_gun'una gore SLA'yi asip asmadigina bakmaksizin) cezalandirir -
    # yani deadline "bir sonraki slot" gibi davranilir, oysa gercek hedef_teslim_gun
    # (route_matrix'ten) 1-2 GUN (24-48 saat). Aggregate ertelenen_talep degiskeni
    # backlog'un "yasini" (ne zaman olusturuldugunu) tutmadigi icin dogru bir
    # deadline karsilastirmasi bu modelde yapilamiyor - bu yuzden SLA maliyeti
    # burada muhtemelen GERCEK PDF kuralina gore OLDUGUNDAN FAZLA gosteriliyor.
    # ALNS motoru (src/alns_engine.py) bunu talep bazinda gercek deadline'a
    # karsi dogru hesapliyor (bkz. try_insert_path/demand_gun ayrimi) - bu yuzden
    # ALNS'in raporladigi SLA rakami CP-SAT'inkinden cok daha guvenilir kabul edilmeli.
    for (g, s) in zaman_sirali:
        saat_farki = _adim_saat_farki(g, s)
        ceza_katsayi = int(round(sla_cezasi_tl(1.0, saat_farki)))  # birim (1 desi) basina TL
        if ceza_katsayi > 0:
            maliyet_kalemleri.append(ertelenen_talep[(h, g, s)] * ceza_katsayi)

model.Minimize(cp_model.LinearExpr.Sum(maliyet_kalemleri))

# ============================================================================
# 8. ÇÖZÜCÜ
# ============================================================================
solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds = ENV_MAX_TIME
solver.parameters.log_search_progress = ENV_LOG_PROGRESS
solver.parameters.num_search_workers = multiprocessing.cpu_count()

status = solver.Solve(model)

# ============================================================================
# 9. ÇIKTI
# ============================================================================
results_dir = PROJECT_ROOT / "results"
results_dir.mkdir(parents=True, exist_ok=True)
output_txt = results_dir / "optimization_results.txt"
output_csv = results_dir / "optimization_results.csv"
output_xlsx = results_dir / "optimization_results.xlsx"

csv_records = []

with output_txt.open("w", encoding="utf-8") as f:
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        durum = "OPTIMAL (Kusursuz)" if status == cp_model.OPTIMAL else f"UYGUN ({ENV_MAX_TIME}s)"
        f.write(f"{'=' * 80}\nOPTIMIZASYON BASARILI - {durum}\n{'=' * 80}\n\n")
        print(f"OPTIMIZASYON BASARILI - {durum}")

        spot_toplam_maliyet = 0
        toplam_ertelenen_desi = 0
        d_rented_count = 0
        d_spot_count = 0

        for h in hatlar:
            src, dst = h
            for (g, s) in zaman_sirali:
                for a in arac_turleri:
                    p = arac_parametreleri[a]
                    k_adet = solver.Value(kiralik_x[(h, g, s, a)])
                    s_adet = solver.Value(spot_y[(h, g, s, a)])
                    d_rented_count += k_adet
                    d_spot_count += s_adet

                    if k_adet > 0:
                        yuk_a = solver.Value(kiralik_yuk_by_type[(h, g, s, a)])
                        birim_maliyet = _maliyet_katsayisi(h, a, p["rental_hourly"], p["rental_km"])
                        toplam_maliyet = k_adet * birim_maliyet
                        f.write(f"{g} {s} | Kiralik {a} | {src}->{dst} | {yuk_a} | {toplam_maliyet}\n")
                        csv_records.append({
                            "Tarih": g, "Slot": s, "Arac_Tipi": f"Kiralik {a}",
                            "Cikis_TM": src, "Varis_TM": dst, "Arac_Sayisi": k_adet,
                            "Teslim_Edilen_Desi": yuk_a, "Maliyet_TL": toplam_maliyet,
                        })

                    if s_adet > 0:
                        yuk_a = solver.Value(spot_yuk_by_type[(h, g, s, a)])
                        birim_maliyet = _maliyet_katsayisi(h, a, p["spot_hourly"], p["spot_km"])
                        toplam_maliyet = s_adet * birim_maliyet
                        spot_toplam_maliyet += toplam_maliyet
                        f.write(f"{g} {s} | Spot {a} | {src}->{dst} | {yuk_a} | {toplam_maliyet}\n")
                        csv_records.append({
                            "Tarih": g, "Slot": s, "Arac_Tipi": f"Spot {a}",
                            "Cikis_TM": src, "Varis_TM": dst, "Arac_Sayisi": s_adet,
                            "Teslim_Edilen_Desi": yuk_a, "Maliyet_TL": toplam_maliyet,
                        })

                ert = solver.Value(ertelenen_talep[(h, g, s)])
                if ert > 0:
                    # NOT: Ayni yuk, teslim edilene kadar birden fazla slotta "ertelenmis"
                    # olarak gorunebilir - bu toplam kumulatiftir, NIHAI teslim edilemeyen
                    # miktar degildir (o her zaman 0 - bkz. asagidaki nihai_teslim_edilemeyen).
                    toplam_ertelenen_desi += ert
                    f.write(f"{g} {s} | ERTELEME | {src}->{dst} | {ert} | 0\n")

        sla_ceza_toplam = 0
        for h in hatlar:
            for (g, s) in zaman_sirali:
                ert = solver.Value(ertelenen_talep[(h, g, s)])
                if ert > 0:
                    saat_farki = _adim_saat_farki(g, s)
                    sla_ceza_toplam += int(round(sla_cezasi_tl(ert, saat_farki)))

        # Modelde son (gun, slot) icin ertelenen_talep == 0 sert kisiti var (bkz. bolum 5) -
        # yani 7 gunun sonunda TUM talep teslim edilmis olmak ZORUNDA (feasible cozumde
        # bu her zaman 0'dir). Burada bunu ayrica hesaplayip acikca raporluyoruz.
        nihai_teslim_edilemeyen = sum(
            solver.Value(ertelenen_talep[(h, son_gun, son_slot)]) for h in hatlar
        )

        genel_toplam = kiralik_sabit_toplam + spot_toplam_maliyet + sla_ceza_toplam

        ozet = f"""
{'=' * 80}
OZET ISTATISTIKLER (Faz 2 - Saat Bazli)
{'=' * 80}
  Kiralik Arac Sabit Maliyeti : {kiralik_sabit_toplam:>15,.0f} TL
  Spot Arac Maliyeti          : {spot_toplam_maliyet:>15,.0f} TL
  SLA Gecikme Cezasi          : {sla_ceza_toplam:>15,.0f} TL
{'-' * 80}
  TOPLAM MALIYET              : {genel_toplam:>15,.0f} TL
{'=' * 80}
  Nihai Teslim Edilemeyen     : {nihai_teslim_edilemeyen:>15,.0f} desi  (7 gun sonunda - sifir olmasi garantili sert kisittir)
  Kumulatif Erteleme-Ani Yuku : {toplam_ertelenen_desi:>15,.0f} desi  (ayni yuk teslim olana kadar birden fazla slotta sayilir - GECIKME SIDDETI gostergesidir, kaybolan yuk degil)
  Cozucu Suresi               : {solver.WallTime():>15.2f} sn
  Direkt Kiralik Arac Sayisi  : {d_rented_count}
  Direkt Spot Arac Sayisi     : {d_spot_count}
{'=' * 80}
"""
        f.write(ozet)
        print(ozet)
    else:
        hata = f"Cozum bulunamadi! Durum kodu: {status}\n"
        f.write(hata)
        print(hata)

if csv_records:
    pd.DataFrame(csv_records).to_csv(output_csv, index=False, encoding="utf-8-sig")
    print(f"CSV kaydedildi: {output_csv}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Teslim Plani"
    basliklar = ["Tarih", "Slot", "Arac Tipi", "Cikis TM", "Varis TM", "Arac Sayisi", "Teslim Edilen Desi", "Maliyet TL"]
    ws.append(basliklar)
    for hucre in ws[1]:
        hucre.font = Font(bold=True, color="FFFFFF")
        hucre.fill = PatternFill("solid", start_color="4472C4")
        hucre.alignment = Alignment(horizontal="center")
    for rec in csv_records:
        ws.append([
            rec["Tarih"], rec["Slot"], rec["Arac_Tipi"], rec["Cikis_TM"], rec["Varis_TM"],
            rec["Arac_Sayisi"], rec["Teslim_Edilen_Desi"], rec["Maliyet_TL"],
        ])
    for col, width in zip("ABCDEFGH", [12, 8, 16, 14, 14, 12, 16, 12]):
        ws.column_dimensions[col].width = width
    wb.save(output_xlsx)
    print(f"Excel kaydedildi: {output_xlsx}")
else:
    print("Kayit yok - CSV/Excel uretilmedi.")
