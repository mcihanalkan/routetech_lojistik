"""RouteTech Faz 2 — ALNS tabanlı optimizasyon (Stage C).

main.py --engine alns (varsayılan) seçildiğinde bu betiği çalıştırır. Aynı
ROUTETECH_OPTIMIZATION_INPUT / ROUTETECH_MAX_TIME_SECONDS ortam değişkenlerini
okur (src/optimization.py [CP-SAT] ile aynı girdi), ALNS döngüsünü çalıştırır,
aynı çıktı şemasını (results/optimization_results.txt/csv/xlsx) üretir — artık
`Rota_Tipi` sütunuyla konsolidasyonlu (relay) sevkiyatları da gösterir.

Mimari: src/alns_engine.py'deki State/destroy/repair yapılarını kullanır; ana
motor ALNS'tir, CP-SAT (`cpsat_hat_repair`) tek bir hattı tam optimal çözen
küçük-pencere bir repair operatörü olarak devreye girer.
"""

from __future__ import annotations

import json
import os
import sys
import time
from pathlib import Path

import numpy as np
import numpy.random as rnd
import pandas as pd

# alns paketinin SimulatedAnnealing kabul kriteri, cok iyi bir aday cozum
# bulundugunda exp(...) hesabinin sonsuza (inf) tasmasina neden olabiliyor -
# bu ZARARSIZ (davranis dogru: "kesinlikle kabul et" demek), sadece gurultulu
# bir konsol uyarisi. Susturuyoruz.
np.seterr(over="ignore")
from alns import ALNS
from alns.accept import SimulatedAnnealing
from alns.select import RouletteWheel
from alns.stop import MaxRuntime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.cost_model import spot_vehicle_count  # noqa: E402
from src.alns_engine import (  # noqa: E402
    ProblemData,
    State,
    cpsat_hat_repair,
    greedy_repair,
    random_removal,
    tm_overload_removal,
    worst_removal,
)
from src.time_model import build_route_lookup  # noqa: E402

# ============================================================================
# 1. ORTAM DEĞİŞKENLERİ
# ============================================================================
ENV_MAX_TIME = float(os.environ.get("ROUTETECH_MAX_TIME_SECONDS", "300")) # Algoritmanın çalışma süresi, terminalden parametre olarak verilebilir
ENV_VERBOSE = os.environ.get("ROUTETECH_LOG_SEARCH_PROGRESS", "1") == "1" 

DEFAULT_INPUT_JSON = PROJECT_ROOT / "data" / "outputs" / "optimization_input.json" 
INPUT_JSON_PATH = Path(os.environ.get("ROUTETECH_OPTIMIZATION_INPUT", str(DEFAULT_INPUT_JSON)))# Optimizasyon input dosyasının path'i

if not INPUT_JSON_PATH.exists():
    raise FileNotFoundError(
        f"Optimizasyon girdisi bulunamadı: {INPUT_JSON_PATH}. Once `python main.py` calistirilmali."
    )

with INPUT_JSON_PATH.open("r", encoding="utf-8") as f:
    payload = json.load(f)

df_forecast = pd.DataFrame(payload["forecast"]) # Tahmin verileri
df_route_matrix = pd.DataFrame(payload["route_matrix"]) # mesafeler
df_vehicle_costs = pd.DataFrame(payload["vehicle_costs"]) # araç parametreleri
df_rental_limits = pd.DataFrame(payload["rental_limits"]) # Kiralık stoklar
df_handling_capacity = pd.DataFrame(payload["handling_capacity"]) # Elleçleme kapasiteleri
df_tir_capacity = pd.DataFrame(payload["tir_capacity"]) # Tır kapasiteleri

if df_forecast.empty:
    raise ValueError(f"Forecast payload bos: {INPUT_JSON_PATH}")

route_lookup = build_route_lookup(df_route_matrix) # veri tablosunu hızlı sorgulanabilir bir dict'e çevirir.

arac_turleri = df_vehicle_costs["vehicle_type"].tolist() 
tir_arac_turu = next(
    (v for v in arac_turleri if str(v).strip().casefold() in ("tır", "tir")), None
)

# Araç parametrelerine hızlı erişim için bir sözlük oluşturma.
arac_parametreleri = {}
for row in df_vehicle_costs.itertuples():
    arac_parametreleri[row.vehicle_type] = {
        "kapasite_desi": float(row.capacity),
        "rental_hourly": float(row.rental_hourly),
        "rental_km": float(row.rental_km),
        "spot_hourly": float(row.spot_hourly),
        "spot_km": float(row.spot_km),
    }
# "(hat,gun) -> kiralik_stok" şeklinde bir dict.
kiralik_stok_gunluk: dict = {}
for row in df_rental_limits.itertuples():
    hat = (row.source, row.destination)
    key = (hat, row.vehicle_type)
    kiralik_stok_gunluk[key] = kiralik_stok_gunluk.get(key, 0) + int(row.vehicle_count)

handling_capacity = {row.center: float(row.capacity) for row in df_handling_capacity.itertuples()}
tir_capacity = {row.center: float(row.capacity) for row in df_tir_capacity.itertuples()}

# ============================================================================
# 2. TALEP / HAT / GÜN ÇIKARIMI
# ============================================================================
df_forecast = df_forecast.copy()
df_forecast["date"] = pd.to_datetime(df_forecast["date"])
df_forecast["gun_key"] = df_forecast["date"].dt.strftime("%Y-%m-%d")
df_forecast["slot"] = df_forecast["slot"].astype(str) # 09:00 veya 17:00

gunler = sorted(df_forecast["gun_key"].unique())
merkezler = sorted(set(handling_capacity) | set(tir_capacity))

# (hat,gun,slot) -> desi şeklinde bir dict. slot'tan kasıt 09:00, 17:00
talep_verisi: dict = {}
for row in df_forecast.itertuples():
    hat = (row.source, row.destination)
    key = (hat, row.gun_key, row.slot)
    talep_verisi[key] = talep_verisi.get(key, 0.0) + max(0.0, float(row.recommended_demand))

# Desi'si 0 olan talepleri göz ardı etmek için yeni liste.
demands = [
    (hat, gun, slot, round(desi))
    for (hat, gun, slot), desi in talep_verisi.items()
    if round(desi) > 0
]

print(f"ALNS: {len(demands):,} talep parcasi, {len(gunler)} gun, {len(merkezler)} TM.")

# ============================================================================
# 3. PROBLEM VERISI VE BASLANGIC COZUMU
# ============================================================================
# Lojistik problemimize ait tüm optimizasyon dataları.
data = ProblemData(
    route_lookup=route_lookup,
    arac_turleri=arac_turleri,
    arac_parametreleri=arac_parametreleri,
    kiralik_stok_gunluk=kiralik_stok_gunluk,
    handling_capacity=handling_capacity,
    tir_capacity=tir_capacity,
    tir_arac_turu=tir_arac_turu,
    gunler=gunler,
    merkezler=merkezler,
    demands=demands,
)

rng = rnd.default_rng(42) # Neden 42 seedini veriyoruz generator'ü oluşturmak için?

initial_state = State(data) 
initial_state = greedy_repair(initial_state, rng)
initial_obj = initial_state.objective()
print(f"Baslangic (greedy) cozum maliyeti: {initial_obj:,.0f} TL")

# ============================================================================
# 4. ALNS KURULUMU
# ============================================================================
alns = ALNS(rng)
alns.add_destroy_operator(random_removal, "random_removal")
alns.add_destroy_operator(worst_removal, "worst_removal")
alns.add_destroy_operator(tm_overload_removal, "tm_overload_removal")
alns.add_repair_operator(greedy_repair, "greedy_repair")
alns.add_repair_operator(cpsat_hat_repair, "cpsat_hat_repair")

select = RouletteWheel(scores=[25, 5, 2, 0.5], decay=0.8, num_destroy=3, num_repair=2)
# num_iters tahmini: cpsat_hat_repair en fazla 5 sn, greedy cok daha hizli -
# ortalama ~2 sn/iterasyon varsayimi (kaba, autofit sicaklik egrisini olceklemek icin yeterli).
tahmini_iterasyon = max(20, int(ENV_MAX_TIME / 2))
accept = SimulatedAnnealing.autofit(
    init_obj=initial_obj, worse=0.05, accept_prob=0.5, num_iters=tahmini_iterasyon
)
stop = MaxRuntime(ENV_MAX_TIME)

# Her yeni en iyi cozum bulundugunda anlik olarak yazdir - onceki CP-SAT'in
# ayrintili arama logunun ALNS'teki esdegeri (bkz. sohbet gecmisi: "eskiden her
# saniye/yenisini bulduğunda görünürdü" - ALNS varsayilan motor olduktan sonra
# bu geri bildirim hic eklenmemisti).
_ilerleme_baslangic = time.time()
_ilerleme_sayac = {"n": 0}


def _yeni_en_iyi_bulundu(candidate_state, rng_):
    _ilerleme_sayac["n"] += 1
    gecen_sn = time.time() - _ilerleme_baslangic
    print(f"  [{gecen_sn:7.1f} sn] Yeni en iyi #{_ilerleme_sayac['n']}: {candidate_state.objective():,.0f} TL")


alns.on_best(_yeni_en_iyi_bulundu)

print(f"ALNS calistiriliyor (bütce: {ENV_MAX_TIME:.0f} sn)...")
result = alns.iterate(initial_state, select, accept, stop)
best: State = result.best_state
print(f"ALNS tamamlandi. En iyi maliyet: {best.objective():,.0f} TL "
      f"(baslangica gore {'%.1f' % (100 * (1 - best.objective() / max(1, initial_obj)))}% iyilesme)")

# ---- Kapasite ihlali dogrulamasi (bu oturumun kontrolu icin, kalici test degil) ----
ihlal_sayisi = 0
for tm, cap in data.handling_capacity.items():
    for gun in data.gunler:
        kullanim = best.handling_usage.get((tm, gun), 0.0)
        if kullanim > cap + 1e-6:
            ihlal_sayisi += 1
            print(f"UYARI: elleceleme kapasitesi asildi: {tm} {gun} kullanim={kullanim:.1f} > kap={cap:.1f}")
for tm, cap in data.tir_capacity.items():
    for gun in data.gunler:
        kullanim = best.tir_usage.get((tm, gun), 0) + data.fixed_kiralik_tir_usage.get((tm, gun), 0)
        if kullanim > cap + 1e-6:
            ihlal_sayisi += 1
            print(f"UYARI: tir kapasitesi asildi: {tm} {gun} kullanim={kullanim} > kap={cap}")
print(f"Kapasite dogrulamasi: {ihlal_sayisi} ihlal bulundu (0 beklenir).")

# ============================================================================
# 5. ÇIKTI — optimization.py ile AYNI ŞEMA (+ Rota_Tipi)
# ============================================================================
results_dir = PROJECT_ROOT / "results"
results_dir.mkdir(parents=True, exist_ok=True)
output_txt = results_dir / "optimization_results.txt"
output_csv = results_dir / "optimization_results.csv"
output_xlsx = results_dir / "optimization_results.xlsx"

def _bucket_key(leg):
    return (leg.src, leg.dst, leg.gun, leg.slot, leg.arac_turu, leg.is_kiralik)


# Bacak (leg-dispatch) basina GERCEK toplam desi - birden fazla talep parcasi
# ayni fiziksel sevkiyati (ayni TM-cift/gun/slot/arac_turu) paylasabilir; tek
# tek assignment'in kendi desi'sinden arac sayisi hesaplamak (eski yontem)
# yaniltici oluyordu (her satir "1 arac" gosterip, ayni bacaga kac FARKLI talep
# parcasinin bindigi gorunmuyordu). Once TUM bacaklarin gercek toplam desisini
# topluyoruz, sonra arac sayisini bundan turetiyoruz.
bucket_toplam_desi: dict = {}
for a in best.assignments:
    for leg in a.legs:
        key = _bucket_key(leg)
        bucket_toplam_desi[key] = bucket_toplam_desi.get(key, 0.0) + a.desi


def _bacak_arac_sayisi(leg) -> int:
    if leg.is_kiralik:
        return 1
    kap = arac_parametreleri[leg.arac_turu]["kapasite_desi"]
    return spot_vehicle_count(bucket_toplam_desi[_bucket_key(leg)], kap, 10 ** 9)


csv_records = []
for a in best.assignments:
    nihai_kaynak, nihai_varis = a.demand_hat
    if len(a.legs) == 1:
        leg = a.legs[0]
        rota_tipi = "Direkt" if leg.gun == a.demand_gun and leg.slot == a.demand_slot else "Direkt (Ertelenmis)"
        arac_tipi = ("Kiralik " if leg.is_kiralik else "Spot ") + leg.arac_turu
        csv_records.append({
            "Tarih": leg.gun, "Slot": leg.slot, "Arac_Tipi": arac_tipi,
            "Cikis_TM": leg.src, "Varis_TM": leg.dst,
            "Nihai_Kaynak": nihai_kaynak, "Nihai_Varis": nihai_varis,
            "Bacaktaki_Arac_Sayisi": _bacak_arac_sayisi(leg),
            "Bu_Talebin_Desisi": round(a.desi, 2),
            "Bacak_Toplam_Desi": round(bucket_toplam_desi[_bucket_key(leg)], 2),
            "Maliyet_TL": round(a.vehicle_cost, 2),
            "Rota_Tipi": rota_tipi, "Talep_Tarihi": a.demand_gun, "Talep_Slotu": a.demand_slot,
        })
    else:
        # Konsolidasyon: bacağın kendi uçları (Cikis_TM/Varis_TM) ile talebin GERÇEK
        # nihai varışı (Nihai_Varis) farklı olabilir — örn. Kocaeli->Eskişehir bacağı,
        # aslında Eskişehir üzerinden Isparta'ya giden bir yükü taşıyor olabilir.
        # Bu ayrım olmadan konsolidasyon anlamsız/gereksiz göründüğü için eklendi.
        ara_duraklar = " -> ".join(leg.dst for leg in a.legs[:-1])  # tum aktarma noktalari (1 ya da 2)
        for i, leg in enumerate(a.legs):
            arac_tipi = ("Kiralik " if leg.is_kiralik else "Spot ") + leg.arac_turu
            csv_records.append({
                "Tarih": leg.gun, "Slot": leg.slot, "Arac_Tipi": arac_tipi,
                "Cikis_TM": leg.src, "Varis_TM": leg.dst,
                "Nihai_Kaynak": nihai_kaynak, "Nihai_Varis": nihai_varis,
                "Bacaktaki_Arac_Sayisi": _bacak_arac_sayisi(leg),
                "Bu_Talebin_Desisi": round(a.desi, 2),
                "Bacak_Toplam_Desi": round(bucket_toplam_desi[_bucket_key(leg)], 2),
                "Maliyet_TL": round(a.vehicle_cost, 2) if i == len(a.legs) - 1 else 0,
                "Rota_Tipi": f"Konsolidasyon {i + 1}/{len(a.legs)} (via {ara_duraklar}, nihai varis: {nihai_varis})",
                "Talep_Tarihi": a.demand_gun, "Talep_Slotu": a.demand_slot,
            })

csv_records.sort(key=lambda r: (r["Tarih"], r["Slot"], r["Cikis_TM"], r["Varis_TM"], r["Talep_Tarihi"], r["Talep_Slotu"]))

# ---- Kapasite kullanim ozeti (elleceleme/tir nasil dahil edildigini GORUNUR kilmak icin) ----
capacity_records = []
for tm, cap in sorted(data.handling_capacity.items()):
    for gun in data.gunler:
        kullanim = best.handling_usage.get((tm, gun), 0.0)
        capacity_records.append({
            "TM": tm, "Tarih": gun, "Tur": "Ellecleme (desi)",
            "Kullanim": round(kullanim, 1), "Kapasite": cap,
            "Doluluk_Yuzde": round(100 * kullanim / cap, 1) if cap else 0,
        })
if data.tir_arac_turu is not None:
    for tm, cap in sorted(data.tir_capacity.items()):
        for gun in data.gunler:
            kullanim = best.tir_usage.get((tm, gun), 0) + data.fixed_kiralik_tir_usage.get((tm, gun), 0)
            capacity_records.append({
                "TM": tm, "Tarih": gun, "Tur": "Tir (adet)",
                "Kullanim": kullanim, "Kapasite": cap,
                "Doluluk_Yuzde": round(100 * kullanim / cap, 1) if cap else 0,
            })

# ---- Arac sevkiyat ozeti (her fiziksel bacak icin TEK satir, dogru arac sayisiyla) ----
dispatch_records = []
for key, toplam_desi in sorted(bucket_toplam_desi.items()):
    src, dst, gun, slot, arac_turu, is_kiralik = key
    kap = arac_parametreleri[arac_turu]["kapasite_desi"]
    adet = 1 if is_kiralik else spot_vehicle_count(toplam_desi, kap, 10 ** 9)
    dispatch_records.append({
        "Tarih": gun, "Slot": slot, "Cikis_TM": src, "Varis_TM": dst,
        "Arac_Tipi": ("Kiralik " if is_kiralik else "Spot ") + arac_turu,
        "Arac_Sayisi": adet, "Toplam_Desi": round(toplam_desi, 2), "Kapasite": kap,
        "Doluluk_Yuzde": round(100 * toplam_desi / (adet * kap), 1) if adet else 0,
    })

kiralik_sabit_toplam = best._fixed_kiralik_cost
spot_toplam_maliyet = sum(a.vehicle_cost for a in best.assignments)
sla_ceza_toplam = sum(a.sla_cost for a in best.assignments)
genel_toplam = best.objective()
konsolidasyon_sayisi = sum(1 for a in best.assignments if len(a.legs) > 1)

ozet = f"""
{'=' * 80}
OZET ISTATISTIKLER (Faz 2 - ALNS, saat bazli, konsolidasyon destekli)
{'=' * 80}
  Kiralik Arac Sabit Maliyeti : {kiralik_sabit_toplam:>15,.0f} TL
  Spot Arac Maliyeti          : {spot_toplam_maliyet:>15,.0f} TL
  SLA Gecikme Cezasi          : {sla_ceza_toplam:>15,.0f} TL
{'-' * 80}
  TOPLAM MALIYET              : {genel_toplam:>15,.0f} TL
{'=' * 80}
  Baslangic (greedy) maliyet  : {initial_obj:>15,.0f} TL
  Toplam atama sayisi         : {len(best.assignments)}
  Konsolidasyonlu (relay) sayi: {konsolidasyon_sayisi}
  Karsilanamayan (garanti 0)  : {sum(x[3] for x in best.unassigned):>15,.2f} desi
{'=' * 80}
"""
print(ozet)
output_txt.write_text(ozet, encoding="utf-8")

output_dispatch_csv = results_dir / "vehicle_dispatch_summary.csv"
output_capacity_csv = results_dir / "capacity_utilization.csv"

if csv_records:
    pd.DataFrame(csv_records).to_csv(output_csv, index=False, encoding="utf-8-sig")
    pd.DataFrame(dispatch_records).to_csv(output_dispatch_csv, index=False, encoding="utf-8-sig")
    pd.DataFrame(capacity_records).to_csv(output_capacity_csv, index=False, encoding="utf-8-sig")
    print(f"CSV kaydedildi: {output_csv}")
    print(f"Arac sevkiyat ozeti: {output_dispatch_csv}")
    print(f"Kapasite kullanim ozeti: {output_capacity_csv}")

    def _yaz_sayfa(ws, basliklar, satirlar, genislikler):
        ws.append(basliklar)
        for hucre in ws[1]:
            hucre.font = Font(bold=True, color="FFFFFF")
            hucre.fill = PatternFill("solid", start_color="4472C4")
            hucre.alignment = Alignment(horizontal="center")
        for row in satirlar:
            ws.append(row)
        for idx, width in enumerate(genislikler, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Teslim Plani (Talep Bazli)"
    _yaz_sayfa(
        ws1,
        ["Tarih", "Slot", "Arac Tipi", "Cikis TM", "Varis TM", "Nihai Kaynak", "Nihai Varis",
         "Bacaktaki Arac Sayisi", "Bu Talebin Desisi", "Bacak Toplam Desi", "Maliyet TL",
         "Rota Tipi", "Talep Tarihi", "Talep Slotu"],
        [[rec["Tarih"], rec["Slot"], rec["Arac_Tipi"], rec["Cikis_TM"], rec["Varis_TM"],
          rec["Nihai_Kaynak"], rec["Nihai_Varis"],
          rec["Bacaktaki_Arac_Sayisi"], rec["Bu_Talebin_Desisi"], rec["Bacak_Toplam_Desi"],
          rec["Maliyet_TL"], rec["Rota_Tipi"], rec["Talep_Tarihi"], rec["Talep_Slotu"]]
         for rec in csv_records],
        [12, 8, 16, 14, 14, 14, 14, 14, 14, 14, 12, 30, 14, 10],
    )

    ws2 = wb.create_sheet("Arac Sevkiyat Ozeti")
    _yaz_sayfa(
        ws2,
        ["Tarih", "Slot", "Cikis TM", "Varis TM", "Arac Tipi", "Arac Sayisi",
         "Toplam Desi", "Kapasite", "Doluluk %"],
        [[r["Tarih"], r["Slot"], r["Cikis_TM"], r["Varis_TM"], r["Arac_Tipi"],
          r["Arac_Sayisi"], r["Toplam_Desi"], r["Kapasite"], r["Doluluk_Yuzde"]]
         for r in dispatch_records],
        [12, 8, 14, 14, 16, 12, 12, 10, 10],
    )

    ws3 = wb.create_sheet("TM Kapasite Kullanimi")
    _yaz_sayfa(
        ws3,
        ["TM", "Tarih", "Tur", "Kullanim", "Kapasite", "Doluluk %"],
        [[r["TM"], r["Tarih"], r["Tur"], r["Kullanim"], r["Kapasite"], r["Doluluk_Yuzde"]]
         for r in capacity_records],
        [14, 12, 18, 12, 12, 10],
    )

    wb.save(output_xlsx)
    print(f"Excel kaydedildi: {output_xlsx} (3 sayfa: Teslim Plani, Arac Sevkiyat Ozeti, TM Kapasite Kullanimi)")
else:
    print("Kayit yok - CSV/Excel uretilmedi.")
