"""
run_forecast.py — Teknofest 2026 Tahmin Çalıştırıcı
=====================================================
Kullanım:
    python run_forecast.py

Giriş  : teknofest26_gelismis.xlsx (slot bazlı: 09:00 / 17:00)
Çıkış  : ortools_payload.csv / .xlsx (slot sütunlu, OR-Tools/ALNS motoruna girdi)
         İsteğe bağlı: alns_payload.json (debug için)

Mimari (İKİ MODEL — Direct Forecasting, 09:00 ve 17:00 ayrı ayrı):
    load_dataset()              → Wide format: (rota, tarih) başına iki hedef
                                   sütun (toplam_desi_0900 / toplam_desi_1700)
    DemandForecaster.fit()      → Her slot için AYRI CatBoost modeli
                                   (MultiQuantile: q10/q50/q90), diğer slot
                                   "sibling_target_column" olarak feature'a girer
    DemandForecaster.predict()  → List[Dict] (tarih, kaynak_tm, varis_tm, slot,
                                   q10, q50, q90) — her model kendi listesini üretir
    UncertaintyBand.to_ortools_dataframe(slot_key="slot") → slot-aware DataFrame;
                                   aynı (date, source, destination) için 09:00 ve
                                   17:00 ayrı satırlar olarak yer alır
    UncertaintyBand.to_alns_payload()      → ALNS formatı (risk_class,
                                   safety_buffer, risk_summary_by_slot, ...)
"""

import json
import logging
import sys
import pandas as pd
import numpy as np
from pathlib import Path
from typing import List, Dict, Any

# Proje modülleri — src/ altında
sys.path.insert(0, str(Path(__file__).parent))
from src.forecasters import DemandForecaster
from src.uncertainty import (
    UncertaintyBand,
    derive_risk_params_from_data,
    combine_slot_bands,
    DYNAMIC_TAU_BASE,
)

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sabitler
# ---------------------------------------------------------------------------

# Proje kökü: run_forecast.py — src/predict_model/ içinde
_HERE          = Path(__file__).resolve().parent          # src/predict_model/
_PROJECT_ROOT  = _HERE.parent.parent                      # routetech_lojistik/

DATA_PATH       = str(_PROJECT_ROOT / "data" / "raw" / "teknofest26_gelismis.xlsx")
MODEL_FILE_PATH_0900 = str(_PROJECT_ROOT / "src" / "predict_model" / "trained_demand_model_0900.joblib")
MODEL_FILE_PATH_1700 = str(_PROJECT_ROOT / "src" / "predict_model" / "trained_demand_model_1700.joblib")
# NOT: Veri 2026-01-01 → 2026-06-28 aralığını kapsıyor (179 gün).
# Gerçek tahmin penceresi — veri sonrası ilk 7 gün (PDF'in "gelecek 7 günlük
# tahmin" mantığına uyar). PREDICT_START artık verinin dışında kaldığı için
# build_predict_grid()'deki buffer mantığı gerçek anlamda devreye giriyor:
# buf_start = PREDICT_START - 35 gün = 2026-05-25, buffer de
# 2026-05-25 → 2026-06-28 arasını (üst sınır zaten < PREDICT_START olduğu
# için otomatik doğru çalışır — smoke-test sırasında eklenen düzeltme
# tam burada işine yarıyor).
PREDICT_START  = "2026-06-29"
PREDICT_END    = "2026-07-05"
OUTPUT_JSON    = str(_HERE / "alns_payload.json")# debug için; ALNS motoru RAM'den alır

DATE_COL    = "tarih"
GROUP_COL   = "rota"          # kaynak_tm → varış_tm kombinasyonu
KAYNAK_COL  = "kaynak_tm"
VARIS_COL   = "varis_tm"

# --- Yeni: iki saat dilimli (09:00 / 17:00) veri için slot boyutu ----------
# teknofest26_gelismis.xlsx her talebi tek satırda, "talep_tamamlanma_saati"
# sütunuyla 09:00 veya 17:00 olarak işaretliyor. load_dataset() artık bu
# üçüncü boyutu (rota × tarih × slot) tam grid'e ekleyip, sonrasında
# (rota, tarih) bazında wide formata pivotluyor: iki paralel hedef sütun
# (TARGET_COL_0900 / TARGET_COL_1700) → Direct Forecasting mimarisinin temeli.
SLOT_COL         = "saat_dilimi"
SLOTS            = ["09:00", "17:00"]
TARGET_COL_0900  = "toplam_desi_0900"
TARGET_COL_1700  = "toplam_desi_1700"

# --- Uyarlanabilir lag seçimi (optimize.py ile BİREBİR AYNI eşikler) --------
# lag_21 / lag_30, her rota için ilk N günü NaN yapıp feature matrix'ten
# düşürüyor (max_lag × rota_sayısı kadar satır kaybı). Küçük veri setlerinde
# bu kayıp sinyale değmiyor; veri büyüdükçe otomatik devreye girsinler diye
# eşik değerine bağlandı. Bu değerler optimize.py'dekiyle AYNI kalmalı —
# aksi halde optimize.py'nin bulduğu hiperparametreler farklı bir feature
# setine göre tuned olur.
LAG_21_MIN_ROWS = 15_000
LAG_30_MIN_ROWS = 20_000


def select_lags(n_real_rows: int) -> list:
    """Veri büyüklüğüne göre lag_21/lag_30'u otomatik ekler/çıkarır — optimize.py ile birebir aynı mantık."""
    lags = [1, 7, 14]
    if n_real_rows >= LAG_21_MIN_ROWS:
        lags.append(21)
    if n_real_rows >= LAG_30_MIN_ROWS:
        lags.append(30)
    return lags


# ---------------------------------------------------------------------------
# 1. Veri Hazırlama
# ---------------------------------------------------------------------------

def _normalize_slot(value: str) -> str:
    """'9:00' / '09:00' / '17:00' gibi varyantları kanonik 'HH:MM' formatına indirger."""
    s = str(value).strip()
    try:
        h, m = s.split(":")
        return f"{int(h):02d}:{int(m):02d}"
    except ValueError:
        # Beklenmeyen bir format gelirse olduğu gibi bırak (aşağıda validasyon uyaracaktır)
        return s


def load_dataset(path: str) -> pd.DataFrame:
    """
    Excel → DemandForecaster'ın beklediği wide formata dönüştür.

    DemandForecaster group_column olarak tek bir sütun bekliyor.
    Grup = kaynak_tm + varış_tm kombinasyonu → 'rota' sütunu.

    ÖNEMLİ — teknofest26_gelismis.xlsx artık slot bazlı (talep_tamamlanma_saati):
    her talep satırı 09:00 veya 17:00 olarak işaretli, aynı (rota, tarih) için
    iki ayrı satır olabiliyor. Bu fonksiyon:
      1. rota × tarih × slot için TAM grid kurar (eksikler 0 — bkz. aşağıdaki not)
      2. wide formata pivotlar → tek satır = (rota, tarih), iki hedef sütun
         (toplam_desi_0900 / toplam_desi_1700)

    Eksikleri neden 0 ile dolduruyoruz (NaN değil):
      Veride 17:00 dolu iken 09:00'in boş olması çok yaygın (17.281 satır),
      tersi nadir (1.069 satır). Bu, eksik slotların veri hatası değil,
      "o slotta gerçekten talep oluşmadı" anlamına geldiğini gösteriyor.
      Dolayısıyla lag/rolling feature'ların süreksiz NaN zincirine
      bölünmemesi için eksikler 0.0 ile doldurulur (eski Desi_talep.xlsx
      sürecindeki mantıkla birebir aynı).

    NOT (Kocaeli): Kocaeli veri setinde hâlâ sadece kaynak olarak geçiyor,
      hiçbir zaman varış değil. Bu, hub/graph feature'larının (hub_in_vol,
      neighbor_vol_2nd_order vb.) varış bazlı hesaplarında Kocaeli için
      sürekli 0 üretecek — hata değil, ama yorumlarken şaşırmamak için not.

    NOT (talep_id): Bu sütun geçmiş gerçek verinin kendi kimliği; bizim
      üreteceğimiz tahmin ID'siyle (D00001 formatı) karışmaması için
      hiçbir feature'a dahil edilmez, bilinçli olarak seçilmez/atılır.

    Sütun eşleştirme stratejisi (farklı dataset varyantlarına dayanıklılık):
      1. Bilinen Türkçe/İngilizce sütun adlarına isim bazlı bak
      2. Bulunamazsa pozisyon bazlı fallback (uyarı logla)
    """
    # NOT: TARGET_COL artık modül seviyesinde bir sabit değil (bkz. Sabitler
    # bölümü) — burada sadece bu fonksiyonun İÇİNDE, pivot'tan önceki ham/tekil
    # "Toplam Desi" sütununu adlandırmak için yerel bir isim kullanıyoruz.
    # Pivot sonrası bu sütun zaten TARGET_COL_0900 / TARGET_COL_1700'e ayrılıyor.
    _RAW_TARGET_COL = "desi_hacmi"

    df = pd.read_excel(path) if path.endswith(".xlsx") else pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]

    # Bilinen olası sütun adları (eski Dataset A Türkçe + yeni teknofest26_gelismis + İngilizce varyantlar)
    _KAYNAK_CANDIDATES = ["Çıkış Transfer Merkezi", "kaynak_tm", "cikis", "çıkış", "source", "origin", "from"]
    _VARIS_CANDIDATES  = ["Varış Transfer Merkezi", "varis_tm",  "varis", "varış", "destination", "dest", "to"]
    _DATE_CANDIDATES   = ["Tarih", "tarih", "date", "Date", "tarih_"]
    _TARGET_CANDIDATES = ["Toplam Desi", "toplam_desi", "desi_hacmi", "desi", "demand", "talep"]
    _SLOT_CANDIDATES   = ["talep_tamamlanma_saati", "saat_dilimi", "slot", "saat", "time_slot"]

    def _find_col(candidates: list, col_idx: int, label: str) -> str:
        """Sütun adını isim bazlı ara, bulamazsan pozisyon bazlı fallback."""
        cols_lower = {c.lower().strip(): c for c in df.columns}
        for cand in candidates:
            if cand in df.columns:
                return cand
            if cand.lower() in cols_lower:
                return cols_lower[cand.lower()]
        # Fallback: pozisyon bazlı
        fallback = df.columns[col_idx]
        logger.warning(
            f"⚠️  '{label}' sütunu isim bazlı bulunamadı → "
            f"pozisyon [{col_idx}] kullanılıyor: '{fallback}'\n"
            f"   Beklenen adlardan biri: {candidates}"
        )
        return fallback

    kaynak_col_raw = _find_col(_KAYNAK_CANDIDATES, 0, "kaynak_tm")
    varis_col_raw  = _find_col(_VARIS_CANDIDATES,  1, "varis_tm")
    date_col_raw   = _find_col(_DATE_CANDIDATES,   2, "tarih")
    target_col_raw = _find_col(_TARGET_CANDIDATES, 3, "desi_hacmi")
    slot_col_raw   = _find_col(_SLOT_CANDIDATES,   4, "saat_dilimi")

    df = df.rename(columns={
        kaynak_col_raw: KAYNAK_COL,
        varis_col_raw:  VARIS_COL,
        date_col_raw:   DATE_COL,
        target_col_raw: _RAW_TARGET_COL,
        slot_col_raw:   SLOT_COL,
    })

    # talep_id gibi bizim işimize yaramayan / tahmin ID'siyle karışabilecek
    # sütunları burada bilinçli olarak dışarıda bırakıyoruz — sadece ihtiyaç
    # duyduğumuz 5 sütunu seçiyoruz.
    df = df[[KAYNAK_COL, VARIS_COL, DATE_COL, _RAW_TARGET_COL, SLOT_COL]].copy()

    df[DATE_COL] = pd.to_datetime(df[DATE_COL])
    df[SLOT_COL] = df[SLOT_COL].apply(_normalize_slot)
    df[GROUP_COL] = df[KAYNAK_COL] + " → " + df[VARIS_COL]

    unexpected_slots = set(df[SLOT_COL].unique()) - set(SLOTS)
    if unexpected_slots:
        logger.warning(f"⚠️  Beklenmeyen saat_dilimi değerleri bulundu: {unexpected_slots}")

    # Aynı (rota, tarih, slot) için birden fazla talep satırı varsa topla
    # (teknofest26_gelismis.xlsx'te her satır zaten tek bir talep ama
    # ileride farklı bir dataset birden fazla talep içerebilir).
    df = (
        df.groupby([GROUP_COL, KAYNAK_COL, VARIS_COL, DATE_COL, SLOT_COL], as_index=False)[_RAW_TARGET_COL]
        .sum()
    )

    # --- Tam grid: rota × tarih × slot (3 boyutlu) — eksikler 0 -----------
    all_dates  = pd.date_range(df[DATE_COL].min(), df[DATE_COL].max(), freq="D")
    all_routes = df[GROUP_COL].unique()
    rota_map   = df[[GROUP_COL, KAYNAK_COL, VARIS_COL]].drop_duplicates()

    idx  = pd.MultiIndex.from_product(
        [all_routes, all_dates, SLOTS], names=[GROUP_COL, DATE_COL, SLOT_COL]
    )
    full = pd.DataFrame(index=idx).reset_index()
    full = full.merge(rota_map, on=GROUP_COL, how="left")
    full = full.merge(
        df[[GROUP_COL, DATE_COL, SLOT_COL, _RAW_TARGET_COL]],
        on=[GROUP_COL, DATE_COL, SLOT_COL],
        how="left",
    )
    full[_RAW_TARGET_COL] = full[_RAW_TARGET_COL].fillna(0.0)

    # --- Wide'a pivot: tek satır = (rota, tarih), iki hedef sütun ----------
    wide = full.pivot_table(
        index=[GROUP_COL, DATE_COL, KAYNAK_COL, VARIS_COL],
        columns=SLOT_COL,
        values=_RAW_TARGET_COL,
        aggfunc="sum",
        fill_value=0.0,
    )
    wide.columns = [
        TARGET_COL_0900 if c == "09:00" else TARGET_COL_1700 for c in wide.columns
    ]
    wide = wide.reset_index()
    wide = wide.sort_values([GROUP_COL, DATE_COL]).reset_index(drop=True)

    logger.info(
        f"✅ Veri hazır (wide format): {len(df):,} gerçek kayıt (slot bazlı) | "
        f"{wide[GROUP_COL].nunique()} rota | {wide[DATE_COL].nunique()} gün\n"
        f"   {wide[DATE_COL].min().date()} → {wide[DATE_COL].max().date()}\n"
        f"   Sütunlar: {TARGET_COL_0900}, {TARGET_COL_1700}"
    )
    return wide


def build_predict_grid(full_df: pd.DataFrame) -> pd.DataFrame:
    """
    PREDICT_START → PREDICT_END aralığı için boş tahmin grid'i oluştur.

    ÖNEMLİ: Artık her satırda İKİ hedef sütun birden bulunuyor
    (TARGET_COL_0900 ve TARGET_COL_1700, ikisi de NaN → 0.0). Bu sayede
    grid her iki model için de ortak kullanılabiliyor — forecasters.py'nin
    predict()'i hangi modelse ona göre kendi target_column'unu ayırıyor.

    Context buffer (son 35 gün) eklenerek lag/rolling doğru hesaplanır.
    """
    target_dates = pd.date_range(PREDICT_START, PREDICT_END, freq="D")
    all_routes   = full_df[GROUP_COL].unique()

    rows = []
    for route in all_routes:
        info = full_df[full_df[GROUP_COL] == route][[KAYNAK_COL, VARIS_COL]].iloc[0]
        for d in target_dates:
            rows.append({
                GROUP_COL:       route,
                DATE_COL:        d,
                KAYNAK_COL:      info[KAYNAK_COL],
                VARIS_COL:       info[VARIS_COL],
                TARGET_COL_0900: np.nan,
                TARGET_COL_1700: np.nan,
            })

    pred_df   = pd.DataFrame(rows)
    buf_start = pd.Timestamp(PREDICT_START) - pd.Timedelta(days=35)
    # ÖNEMLİ: Üst sınır da şart. PREDICT_START/END, full_df'in KAPSADIĞI bir
    # aralıksa (örn. smoke test — geçmiş bir haftayı "tahmin" ediyorsak),
    # üst sınır olmadan buffer PREDICT_START-END için de gerçek/dolu satırları
    # içerir. pred_df aynı (rota, tarih) kombinasyonları için NaN satırlar
    # eklediğinde concat sonrası her kombinasyon için İKİ satır oluşur (biri
    # gerçek, biri NaN) — bu da lag/rolling hesaplarını (shift/rolling
    # pencereleri kayar) ve predict()'in tahmin satırlarını ayırt etmesini
    # bozar. Bu yüzden buffer kesinlikle PREDICT_START'tan ÖNCEKİ günlerle
    # sınırlı olmalı — PREDICT_START/END ileride gerçek geleceğe (full_df'in
    # dışına) çekilse bile bu satır genel/doğru kalır.
    buffer    = full_df[
        (full_df[DATE_COL] >= buf_start) & (full_df[DATE_COL] < pd.Timestamp(PREDICT_START))
    ].copy()
    combined  = pd.concat([buffer, pred_df], ignore_index=True)
    combined  = combined.sort_values([GROUP_COL, DATE_COL]).reset_index(drop=True)
    # NaN → 0 (lag kaynağı için), tahmin haftasında model bu değerleri üretecek
    # — iki hedef sütun için de ayrı ayrı uygulanır.
    for col in (TARGET_COL_0900, TARGET_COL_1700):
        combined[col] = combined[col].fillna(0.0)
    return combined


# ---------------------------------------------------------------------------
# 2. Tahmin + ALNS Payload
# ---------------------------------------------------------------------------

def _fit_or_load_forecaster(
    full_df: pd.DataFrame,
    target_column: str,
    sibling_target_column: str,
    model_path: Path,
    label: str,
) -> DemandForecaster:
    """
    Tek bir slot (09:00 ya da 17:00) için model yükle/eğit yardımcı fonksiyonu.
    Model dosyası varsa doğrudan yüklenir (bu durumda select_lags() hiç
    çağrılmaz — eski davranışla birebir aynı korunuyor). Yoksa n_real_rows
    bu slota özgü hesaplanır ve kendi lag seti seçilir (iki model farklı
    veri yoğunluğuna sahip olabileceğinden farklı lag setleriyle kurulabilir).
    """
    if model_path.exists():
        logger.info(f"📦 [{label}] Hazır eğitilmiş model bulundu! Yükleniyor: {model_path.name}")
        try:
            forecaster = DemandForecaster.load_model(str(model_path))
            # Hızlı bir doğrulama: yüklenen modelin feature importance'a
            # erişimi çalışıyor mu kontrol et (predict() ile uyumsuz/bozuk
            # bir .joblib varsa burada patlar). Bu satır, run()'ın ana
            # akışında zaten çağrılan get_feature_importances()'ı ÖNCEDEN
            # tetikleyerek, dosya üretim adımlarına gelmeden önce sorunu
            # yakalayıp fallback'e düşmemizi sağlıyor — aksi halde model
            # yüklendikten sonraki bir adımda (predict/importance) sessizce
            # patlayan hatalar, ortools/xlsx çıktılarının HİÇ üretilmemesine
            # (çünkü run() istisna fırlatıp yarıda kesiliyor) yol açıyordu.
            forecaster.get_feature_importances()
            return forecaster
        except Exception as exc:
            logger.warning(
                f"⚠️ [{label}] Hazır model yüklenemedi/uyumsuz ({exc!r}). "
                f"Model sıfırdan eğitilerek devam edilecek — çıktı dosyaları "
                f"yine de üretilecek."
            )
            # Bozuk/uyumsuz model dosyasını temizle ki bir sonraki çalıştırmada
            # tekrar aynı hatayla karşılaşılmasın.
            try:
                model_path.unlink()
            except OSError:
                pass

    logger.info(f"⚠️ [{label}] Hazır model dosyası bulunamadı. Model sıfırdan eğitiliyor...")

    n_real_rows = int((full_df[target_column] > 0).sum())
    lags = select_lags(n_real_rows)
    logger.info(
        f"   [{label}] Gerçek kayıt: {n_real_rows:,} → lag'ler: {lags} "
        f"(eşikler: lag_21≥{LAG_21_MIN_ROWS:,}, lag_30≥{LAG_30_MIN_ROWS:,})"
    )

    clip_mult = 5.0 if "0900" in target_column else 3.0

    forecaster = DemandForecaster(
        target_column          = target_column,
        sibling_target_column  = sibling_target_column,
        date_column             = DATE_COL,
        group_column            = GROUP_COL,
        train_test_split        = 0.85,
        forecast_horizon        = 7,
        lags                    = lags,   # veri büyüklüğüne göre uyarlanır — bkz. select_lags()
        rolling_windows         = [7, 14],
        outlier_clip_multiplier = clip_mult,
        logging_enabled         = True,
        random_state            = 42,
        campaign_release_alpha   = 2.5,   # alpha kanıtlanmış şekilde etkisiz (accumulated_campaign_eve_days /
                                           # days_since_campaign_end ham kolonları zaten X'te olduğu için
                                           # CatBoost'a redundant geliyor) → sınıf varsayılanında bırakıldı.
        campaign_max_release_days = 7,    # sweep sonucu kazanan (bkz. özet, decision_regret=1155.96)
    )
    forecaster.fit(full_df)

    # Eğitilen modeli gelecekte kullanmak üzere kaydet (.joblib dosyası olarak)
    forecaster.save_model(str(model_path))
    return forecaster


def run(save_json: bool = True) -> "pd.DataFrame":
    """
    Ana akış (İKİ MODEL — Direct Forecasting: 09:00 ve 17:00 ayrı modeller):
      1. Veri yükle (tek çağrı — full_df her iki slotu da wide formatta içeriyor)
      2. Her slot için ayrı DemandForecaster.fit() (ya da hazır modeli yükle)
      3. Ortak predict_grid ile her iki model için predict() → 2 ayrı List[Dict]
      4. İki listeyi birleştir (predict()'ten gelen "slot" alanı korunuyor)
      5. UncertaintyBand.to_ortools_dataframe(..., slot_key="slot") — artık
         slot-aware; aynı (date, source, destination) için 09:00 ve 17:00
         AYRI satırlar olarak DataFrame'de yer alıyor.
      6. İsteğe bağlı ALNS JSON payload + CSV/Excel çıktıları kaydedilir.
    """
    logger.info("=" * 60)
    logger.info("🚀 Teknofest 2026 — Tahmin Motoru (09:00 & 17:00)")
    logger.info("=" * 60)

    # --- 1. Veri (tek çağrı, her iki slot da içinde) ---
    full_df = load_dataset(DATA_PATH)

    # --- 2. İki model: 09:00 ve 17:00 ayrı ayrı eğit/yükle ---
    forecaster_0900 = _fit_or_load_forecaster(
        full_df, TARGET_COL_0900, TARGET_COL_1700, Path(MODEL_FILE_PATH_0900), "09:00"
    )
    forecaster_1700 = _fit_or_load_forecaster(
        full_df, TARGET_COL_1700, TARGET_COL_0900, Path(MODEL_FILE_PATH_1700), "17:00"
    )

    # --- Feature Importance (her model kendi feature önceliğini gösterir —
    #     örn. 17:00 modeli muhtemelen toplam_desi_0900'e yüksek önem verecek) ---
    importances_0900 = forecaster_0900.get_feature_importances()
    print("\n\U0001f525 [09:00] En Önemli 10 Feature:")
    print(importances_0900.head(10))

    importances_1700 = forecaster_1700.get_feature_importances()
    print("\n\U0001f525 [17:00] En Önemli 10 Feature:")
    print(importances_1700.head(10))

    # --- 3. Ortak predict grid (her iki hedefi birden taşıyor) ---
    predict_grid = build_predict_grid(full_df)

    # --- TEŞHİS 2: 06-29 vs 06-30 feature seviyesi karşılaştırma ---
    _diag_routes = ["Yalova → İstanbul", "Kocaeli → İstanbul", "İstanbul → Yalova"]
    _diag_routes = [r for r in _diag_routes if r in predict_grid[GROUP_COL].unique()]
    if not _diag_routes:
        _diag_routes = [predict_grid[GROUP_COL].unique()[0]]
    for _label, _fc in [("09:00", forecaster_0900), ("17:00", forecaster_1700)]:
        _horizon_mask = (
            (predict_grid[DATE_COL] >= pd.Timestamp(PREDICT_START)) &
            (predict_grid[DATE_COL] <= pd.Timestamp(PREDICT_END))
        )
        _pg = predict_grid.loc[_horizon_mask].copy()
        _pg["_is_predict_row_"] = True
        _combined = _fc._prepend_context_buffer(_pg)
        _combined["_is_predict_row_"] = _combined["_is_predict_row_"].fillna(False)
        _feat = _fc._engineer_features(_combined, drop_na=False)
        _feat = _feat[_feat["_is_predict_row_"] == True]
        _cols = [c for c in [
            "rota", "tarih", "is_closed", "days_since_resumption",
            "accumulated_closed_days", "backlog_release_index",
            f"lag_1_{'0900' if _label=='09:00' else '1700'}",
            f"rolling_mean_7_{'0900' if _label=='09:00' else '1700'}",
            f"rolling_mean_14_{'0900' if _label=='09:00' else '1700'}",
            f"rolling_std_7_{'0900' if _label=='09:00' else '1700'}",
            f"rolling_std_14_{'0900' if _label=='09:00' else '1700'}",
            "hub_out_vol_7d", "hub_in_vol_7d",
        ] if c in _feat.columns]
        print(f"\n🔬 [{_label}] Feature teşhisi — {_diag_routes}:")
        _sub = _feat[_feat["rota"].isin(_diag_routes)][_cols].sort_values(["rota", "tarih"])
        print(_sub.to_string(index=False))

    # ÖNEMLİ: 7 günlük ufuk artık TEK SEFERDE değil, GÜN GÜN (recursive/
    # autoregressive) tahmin ediliyor — her günün q50'si bir sonraki günün
    # lag/rolling feature'larına "gerçekmiş gibi" besleniyor. Bkz.
    # DemandForecaster.predict_sequential() docstring'i.
    horizon_mask_0900 = (
        (predict_grid[DATE_COL] >= pd.Timestamp(PREDICT_START)) &
        (predict_grid[DATE_COL] <= pd.Timestamp(PREDICT_END))
    )
    predict_grid_0900_horizon = predict_grid.loc[horizon_mask_0900].copy()
    preds_0900: List[Dict[str, Any]] = forecaster_0900.predict_sequential(predict_grid_0900_horizon)

    # --- TEŞHİS 3: lag_1 çöküşü gerçek q50'den mi geliyor, yoksa pred_map
    # eşleşme hatasından mı? (bkz. forecasters.py::predict_sequential()
    # içindeki pred_map.get((rota, tarih), 0.0) fallback'i — anahtar
    # tutmazsa sessizce 0.0'a düşer). Burada sadece o günün q50'sini
    # gösteriyoruz; gerçekten ~0 ise bug değil, gerçek bir sinyaldir.
    _diag_df_0900 = pd.DataFrame(preds_0900)
    _diag_df_0900["tarih"] = _diag_df_0900["tarih"].astype(str).str[:10]
    _diag_cols_0900 = ["rota", "tarih", "q50"] + (
        ["q50_base"] if "q50_base" in _diag_df_0900.columns else []
    )
    print(f"\n🔬 [09:00] Gerçek q50 tahminleri — {_diag_routes}:")
    print(
        _diag_df_0900[_diag_df_0900["rota"].isin(_diag_routes)]
        [_diag_cols_0900]
        .sort_values(["rota", "tarih"])
        .to_string(index=False)
    )

    # --- TEŞHİS: recursive vs non-recursive karşılaştırma (geçici) ---
    preds_0900_nonrec = forecaster_0900.predict(predict_grid_0900_horizon)
    _df_diag = pd.DataFrame(preds_0900_nonrec)
    _df_diag["tarih"] = _df_diag["tarih"].astype(str).str[:10]
    logger.info("\n🔬 [09:00] TEŞHİS — recursive(sequential) vs non-recursive günlük toplam:")
    _seq_daily = pd.DataFrame(preds_0900).assign(tarih=lambda d: d["tarih"].astype(str).str[:10]).groupby("tarih")["q50"].sum()
    _nonrec_daily = _df_diag.groupby("tarih")["q50"].sum()
    for d in sorted(set(_seq_daily.index) | set(_nonrec_daily.index)):
        logger.info(f"   {d} | sequential={_seq_daily.get(d,0):>12,.0f} | non-recursive={_nonrec_daily.get(d,0):>12,.0f}")

    # --- [DÜZELTME] 09:00 tahminlerini 17:00 modelinin grid'ine geri yaz ---
    # build_predict_grid() TARGET_COL_0900'ü tahmin ufkunun TAMAMI için
    # 0.0 ile dolduruyor (gelecek, henüz gerçekleşmemiş). Ama
    # _get_drop_columns() (forecasters.py) sibling_target_column'u (17:00
    # modeli için toplam_desi_0900) BİLEREK feature olarak tutuyor — çünkü
    # 17:00 tahmini yapıldığında sabahki 09:00 talebi zaten gerçekleşmiş
    # sayılır, dolayısıyla leakage değildir ve model buna gerçekten önem
    # veriyor (feature importance'ta top-10). predict_grid'i olduğu gibi
    # 17:00'a verirsek model bu meşru feature için her gün "sabah hiç
    # kargo çıkmadı" (0.0) yalanını görür. Bu yüzden 09:00 modelinin
    # ürettiği q50 tahminlerini, 17:00'a geçmeden önce predict_grid'in
    # TARGET_COL_0900 sütununa (sadece tahmin ufku aralığında) yazıyoruz.
    #
    # ⚠️ Tarih formatı uyuşmazlığı sessizce yutulur: preds_0900'daki
    # DATE_COL değeri predict_sequential()'dan "YYYY-MM-DD" string olarak
    # gelir (bkz. _predict_single_batch → date_vals.dt.strftime), predict_grid
    # içindeki DATE_COL ise datetime'dır. Eşleştirmeden önce ikisini de
    # aynı string formatına çekiyoruz; aksi halde .get() hep fallback'e
    # (eski 0.0) düşer ve hata vermeden sessizce yanlış sonuç üretir.
    pred_0900_lookup: Dict[Any, float] = {
        (r[GROUP_COL], str(r[DATE_COL])[:10]): r["q50"] for r in preds_0900
    }
    predict_grid_1700 = predict_grid.copy()
    mask_horizon = (
        (predict_grid_1700[DATE_COL] >= pd.Timestamp(PREDICT_START)) &
        (predict_grid_1700[DATE_COL] <= pd.Timestamp(PREDICT_END))
    )

    def _inject_0900(row):
        key = (row[GROUP_COL], row[DATE_COL].strftime("%Y-%m-%d"))
        return pred_0900_lookup.get(key, row[TARGET_COL_0900])

    predict_grid_1700.loc[mask_horizon, TARGET_COL_0900] = (
        predict_grid_1700.loc[mask_horizon].apply(_inject_0900, axis=1)
    )

    n_injected = int(
        (predict_grid_1700.loc[mask_horizon, TARGET_COL_0900].values
         != predict_grid.loc[mask_horizon, TARGET_COL_0900].values).sum()
    )
    logger.info(
        f"   🔗 [17:00] {n_injected}/{mask_horizon.sum()} satırda "
        f"toplam_desi_0900, 09:00 modelinin TAHMİNİYLE dolduruldu "
        f"(eskiden hep 0.0 sabitti)."
    )

    predict_grid_1700_horizon = predict_grid_1700.loc[mask_horizon].copy()
    preds_1700: List[Dict[str, Any]] = forecaster_1700.predict_sequential(predict_grid_1700_horizon)

    # --- TEŞHİS 3 (17:00 karşılığı): aynı lag_1/pred_map kontrolü ---
    _diag_df_1700 = pd.DataFrame(preds_1700)
    _diag_df_1700["tarih"] = _diag_df_1700["tarih"].astype(str).str[:10]
    _diag_cols_1700 = ["rota", "tarih", "q50"] + (
        ["q50_base"] if "q50_base" in _diag_df_1700.columns else []
    )
    print(f"\n🔬 [17:00] Gerçek q50 tahminleri — {_diag_routes}:")
    print(
        _diag_df_1700[_diag_df_1700["rota"].isin(_diag_routes)]
        [_diag_cols_1700]
        .sort_values(["rota", "tarih"])
        .to_string(index=False)
    )

    # --- TEŞHİS: recursive vs non-recursive karşılaştırma (geçici) ---
    preds_1700_nonrec = forecaster_1700.predict(predict_grid_1700_horizon)
    _df_diag_1700 = pd.DataFrame(preds_1700_nonrec)
    _df_diag_1700["tarih"] = _df_diag_1700["tarih"].astype(str).str[:10]
    logger.info("\n🔬 [17:00] TEŞHİS — recursive(sequential) vs non-recursive günlük toplam:")
    _seq_daily_1700 = pd.DataFrame(preds_1700).assign(tarih=lambda d: d["tarih"].astype(str).str[:10]).groupby("tarih")["q50"].sum()
    _nonrec_daily_1700 = _df_diag_1700.groupby("tarih")["q50"].sum()
    for d in sorted(set(_seq_daily_1700.index) | set(_nonrec_daily_1700.index)):
        logger.info(f"   {d} | sequential={_seq_daily_1700.get(d,0):>12,.0f} | non-recursive={_nonrec_daily_1700.get(d,0):>12,.0f}")

    # Buffer satırlarını çıkar — sadece PREDICT_START/END aralığı kalsın.
    # İki slotun tahminlerini burada birleştiriyoruz; her kayıtta predict()'ten
    # gelen "slot" alanı zaten korunduğu için birleşim sonrası bile hangi
    # kaydın 09:00'a, hangisinin 17:00'a ait olduğu kaybolmuyor.
    target_dates = set(
        pd.date_range(PREDICT_START, PREDICT_END, freq="D").strftime("%Y-%m-%d")
    )
    # [Slot-bazlı türetim] raw_preds_0900/1700 ayrı ayrı tutulur — her slot
    # kendi materiality_floor/tau_base'ini kendi q50 dağılımından türetecek
    # (bkz. derive_risk_params_from_data() ve aşağıdaki 4. adım).
    raw_preds_0900 = [r for r in preds_0900 if str(r[DATE_COL])[:10] in target_dates]
    raw_preds_1700 = [r for r in preds_1700 if str(r[DATE_COL])[:10] in target_dates]
    raw_preds = raw_preds_0900 + raw_preds_1700

    if not raw_preds:
        raise RuntimeError(
            "❌ Tahmin sonucu 0 kayıt döndü (raw_preds boş). Muhtemel neden: "
            "yüklenen model ile mevcut predict_grid uyumsuz (ör. farklı "
            f"tarih aralığı/feature seti). PREDICT_START={PREDICT_START}, "
            f"PREDICT_END={PREDICT_END}. Modeli silip (trained_demand_model_"
            "0900.joblib / _1700.joblib) yeniden çalıştırmayı deneyin — bu "
            "durumda script otomatik olarak sıfırdan eğitim yapacaktır."
        )

    logger.info(
        f"\n✅ Tahmin tamamlandı: {len(raw_preds)} kayıt "
        f"({PREDICT_START} → {PREDICT_END}, "
        f"{full_df[GROUP_COL].nunique()} rota × 7 gün × 2 slot)"
    )

    # -------------------------------------------------------------------
    # uncertainty.py artık slot-aware (bkz. DemandBand.slot, from_json()'daki
    # slot_key, to_ortools_dataframe()'deki zorunlu "slot" sütunu) — bu adım
    # geri açıldı. 09:00 ve 17:00 tahminleri artık aynı (date, source,
    # destination) için AYRI satırlar olarak DataFrame'de yer alıyor.
    # -------------------------------------------------------------------

    if save_json:
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(raw_preds, f, ensure_ascii=False, indent=2, default=str)
        logger.info(f"💾 Ham tahminler (debug, slot alanlı) kaydedildi: {OUTPUT_JSON}")

    print("\n📋 Ham Tahmin Örneği (İlk 5 Kayıt):")
    for r in raw_preds[:5]:
        print(r)

    # --- 4. Slot-Bazlı / Veri-Türetimli Risk Parametreleri + OR-Tools Payload ---
    #
    # ESKİ DAVRANIŞ (sabit materiality_floor=750.0, TEK UncertaintyBand):
    #   09:00 ve 17:00'nin hacim (q50) dağılımları yapısal olarak çok farklı
    #   (bkz. to_ortools_dataframe altındaki slot-bazlı log notu — 09:00
    #   ort. ~517, 17:00 ort. ~3225, ~6 kat fark). Tek bir sabit floor
    #   kullanmak ya 09:00'ı gereğinden fazla bastırır ya da 17:00'de
    #   gerçekten durgun rotaları yeterince bastırmaz.
    #
    # YENİ DAVRANIŞ: her slotun materiality_floor'u KENDİ q50 dağılımından
    # (p25) türetilir (derive_risk_params_from_data — derive_gamma_from_costs
    # ile aynı felsefe: sabit sayı yerine veriden türetilmiş parametre).
    # 09:00 ve 17:00, kendi parametreleriyle AYRI UncertaintyBand
    # örnekleri üzerinden işlenir, sonra combine_slot_bands() ile TEK bir
    # OR-Tools DataFrame'i / ALNS payload'ında birleştirilir (talep_id'ler
    # birleşim sonrası yeniden ve sıralı atanır — bkz. combine_slot_bands
    # docstring'i).
    #
    # [derive_tau=True] tau_base de artık her slotun KENDİ gözlemlenen
    # ortalama U_rel'inden türetiliyor. Global DYNAMIC_TAU_BASE (0.50)
    # tüm filonun ortalama ölçeğine göre kalibre edilmişti — ama 09:00
    # slotu, 17:00'ye göre çok daha düşük hacimli (~6 kat fark, bkz.
    # yukarıdaki not) olduğundan gerçek U_rel tabanı da çok farklı
    # (09:00'da gözlemlenen U_rel ortalaması, sabit 0.50'nin ~10 katına
    # kadar çıkabiliyor). Sabit tau_base bu durumda 09:00 için anlamsız
    # kalıyor: neredeyse her satır dinamik eşiğin üzerine çıkıp yapay
    # şekilde MEDIUM/HIGH'a itiliyor. q10/q90 verilerek derive_tau=True
    # geçildiğinde her slot kendi gerçek U_rel tabanına göre kalibre olur
    # (sıfır-hariç istatistikle — bkz. derive_risk_params_from_data
    # docstring'i, "durgun" satırlar bu ortalamayı şişirmesin diye dışlanır).
    #
    # [Tur 7 DÜZELTMESİ] İlk üretim çalıştırmasında (bu dosyanın önceki
    # sürümü) 09:00 slotu için tau_base=0.872 çıkmıştı — relative_uncertainty
    # artık [0,1) ile sınırlı olduğundan (bkz. uncertainty.py "Tur 6" notu),
    # bu tavan'a (1.0) neredeyse yapışık bir eşikti ve HIGH sınıfını fiilen
    # imkânsız kılıyordu (2023 kayıtta HIGH=0). Kök neden: 09:00 gibi seyrek/
    # durgun bir slotta TİPİK (medyan) ham oran bile zaten çok yüksek
    # (~%491) olduğundan, düz arctan(x) onu direkt tavana taşıyordu. Çözüm:
    # derive_risk_params_from_data artık ayrıca "maape_scale" (o slotun
    # tipik ham oranı) döndürüyor ve bu, UncertaintyBand(maape_scale=...)
    # parametresine aktarılıyor — böylece tau_base tüm slotlarda taşınabilir/
    # sabit kalıyor, gerçek outlier'lar (medyanın kat kat üstü) hâlâ HIGH
    # tetikleyebiliyor.
    params_0900 = derive_risk_params_from_data(
        q50_values=[r.get("q50", 0.0) for r in raw_preds_0900],
        q10_values=[r.get("q10", 0.0) for r in raw_preds_0900],
        q90_values=[r.get("q90", 0.0) for r in raw_preds_0900],
        floor_percentile=25.0,
        derive_tau=True,
    )
    params_1700 = derive_risk_params_from_data(
        q50_values=[r.get("q50", 0.0) for r in raw_preds_1700],
        q10_values=[r.get("q10", 0.0) for r in raw_preds_1700],
        q90_values=[r.get("q90", 0.0) for r in raw_preds_1700],
        floor_percentile=25.0,
        derive_tau=True,
    )
    logger.info(
        f"\n📐 Slot-bazlı türetilmiş risk parametreleri → "
        f"09:00: floor={params_0900['materiality_floor']:.1f}, "
        f"scale={params_0900.get('maape_scale', 1.0):.3f}, "
        f"tau_base={params_0900.get('tau_base', DYNAMIC_TAU_BASE):.3f} | "
        f"17:00: floor={params_1700['materiality_floor']:.1f}, "
        f"scale={params_1700.get('maape_scale', 1.0):.3f}, "
        f"tau_base={params_1700.get('tau_base', DYNAMIC_TAU_BASE):.3f} "
        f"(p25 of q50 + ort. gözlemlenen U_rel, her slotun kendi dağılımından, sıfır-hariç)"
    )

    band_0900 = UncertaintyBand(
        buffer_ratio=0.5,
        logging_enabled=True,
        materiality_floor=params_0900["materiality_floor"],
        tau_base=params_0900.get("tau_base", DYNAMIC_TAU_BASE),
        maape_scale=params_0900.get("maape_scale", 1.0),
    )
    band_1700 = UncertaintyBand(
        buffer_ratio=0.5,
        logging_enabled=True,
        materiality_floor=params_1700["materiality_floor"],
        tau_base=params_1700.get("tau_base", DYNAMIC_TAU_BASE),
        maape_scale=params_1700.get("maape_scale", 1.0),
    )

    # Not: to_ortools_dataframe() yerine doğrudan from_json() çağrılıyor —
    # combine_slot_bands() zaten kendi DataFrame/payload üretimini yapıyor;
    # to_ortools_dataframe()'i burada çağırmak sadece talep_id'leri gereksiz
    # yere (D00001'den) iki kez üretip hemen ardından ezmiş olurdu.
    band_0900.from_json(raw_preds_0900, date_key=DATE_COL, group_key=GROUP_COL, slot_key="slot")
    band_1700.from_json(raw_preds_1700, date_key=DATE_COL, group_key=GROUP_COL, slot_key="slot")

    combined = combine_slot_bands([band_0900, band_1700])
    df_ortools = combined["dataframe"]

    # --- 5. Çıktıları Kaydet (Algoritma ve Jüri İçin) ---

    if save_json:
        payload = combined["payload"]
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        logger.info(f"ALNS payload kaydedildi: {OUTPUT_JSON}")

    # 1. ALGORİTMA İÇİN CSV ÇIKTISI
    OUTPUT_CSV = str(_HERE / "ortools_payload.csv")
    df_ortools.to_csv(OUTPUT_CSV, index=False)
    logger.info(f"💾 OR-Tools (Algoritma) payload kaydedildi: {OUTPUT_CSV}")

    # 2. JÜRİ İÇİN EXCEL ÇIKTISI (Results klasörüne)
    excel_dir = _PROJECT_ROOT / "results"
    excel_dir.mkdir(parents=True, exist_ok=True)
    OUTPUT_EXCEL = str(excel_dir / "ortools_payload.xlsx")
    df_ortools.to_excel(OUTPUT_EXCEL, index=False)
    logger.info(f"📊 Jüri (Raporlama) payload kaydedildi: {OUTPUT_EXCEL}")

    # -------------------------------------------------------------------
    # 3. PDF TESLİM FORMATI — Talep-Tahmini.xlsx
    # -------------------------------------------------------------------
    # df_ortools zaten talep_id sütununu taşıyor (bkz. uncertainty.py::
    # UncertaintyBand.from_json — sıralı D00001... ataması). Burada sadece
    # jürinin resmi TALEP_TAHMI_NI_.xlsx şablonuyla BİREBİR aynı sütun
    # adları / sırası / hücre tipleriyle AYRI bir teslim dosyası
    # üretiyoruz — ortools_payload.csv/xlsx (algoritma girdisi, risk_class
    # vb. dahil zengin format) buna DOKUNULMADAN aynen kalıyor.
    #
    # Şablon (TALEP_TAHMI_NI_.xlsx) sütunları — birebir referans:
    #   A: Talep ID                  → örn. 'D00001'          (metin)
    #   B: Tarih                     → örn. '01.04.2026'       (metin, GG.AA.YYYY)
    #   C: Talep Tamamlama Saati     → örn. 09:00 / 17:00      (datetime.time, hücre formatı 'h:mm')
    #   D: Çıkış Transfer Merkezi    → örn. 'X'                (metin)
    #   E: Varış Transfer Merkezi    → örn. 'Y'                (metin)
    #   F: Tahmin Edilen Desi        → örn. 4.3                (sayı, hücre formatı '0.000')
    from datetime import time as _dt_time

    def _slot_to_time(slot_str: str) -> _dt_time:
        """'09:00' / '17:00' -> datetime.time nesnesi (sablondaki C sutunu tipiyle birebir)."""
        h, m = str(slot_str).split(":")
        return _dt_time(int(h), int(m))

    talep_tahmini_df = df_ortools.rename(columns={
        "talep_id":    "Talep ID",
        "source":      "Çıkış Transfer Merkezi",
        "destination": "Varış Transfer Merkezi",
    }).copy()

    talep_tahmini_df["Tarih"] = pd.to_datetime(talep_tahmini_df["date"]).dt.strftime("%d.%m.%Y")
    talep_tahmini_df["Talep Tamamlama Saati"] = talep_tahmini_df["demand_start_time"].apply(_slot_to_time)
    talep_tahmini_df["Tahmin Edilen Desi"] = talep_tahmini_df["q50"].astype(float)

    # Şablonla BİREBİR aynı sütun sırası
    talep_tahmini_df = talep_tahmini_df[[
        "Talep ID", "Tarih", "Talep Tamamlama Saati",
        "Çıkış Transfer Merkezi", "Varış Transfer Merkezi", "Tahmin Edilen Desi",
    ]]

    # --- Toplam Talep Tahmini (desi) — tüm rota/tarih/saat dilimi toplamı ---
    total_desi_tahmini = float(talep_tahmini_df["Tahmin Edilen Desi"].sum())

    OUTPUT_TALEP_XLSX = str(excel_dir / "Talep-tahmini.xlsx")
    talep_tahmini_df.to_excel(OUTPUT_TALEP_XLSX, index=False)

    # Hücre formatlarını da şablonla birebir eşleştir:
    #   C sütunu (Talep Tamamlama Saati) -> 'h:mm'
    #   F sütunu (Tahmin Edilen Desi)    -> '0.000'
    from openpyxl import load_workbook
    _wb_out = load_workbook(OUTPUT_TALEP_XLSX)
    _ws_out = _wb_out.active
    for _row in _ws_out.iter_rows(min_row=2, max_row=_ws_out.max_row):
        _row[2].number_format = "h:mm"   # C: Talep Tamamlama Saati
        _row[5].number_format = "0.000"  # F: Tahmin Edilen Desi

    # --- Toplam satırı: veri satırlarından bir boşluk sonra, E/F sütunlarına ---
    _total_row_idx = _ws_out.max_row + 2
    _ws_out.cell(row=_total_row_idx, column=5, value="TOPLAM TALEP TAHMİNİ (desi)")
    _total_cell = _ws_out.cell(row=_total_row_idx, column=6, value=total_desi_tahmini)
    _total_cell.number_format = "0.000"

    _wb_out.save(OUTPUT_TALEP_XLSX)

    logger.info(f"📋 Jüri teslim formatı (şablona birebir uygun) kaydedildi: {OUTPUT_TALEP_XLSX}")

    logger.info(
        f"\n{'='*60}\n"
        f"✅ Tamamlandı! (model hazırdan yüklendi ya da sıfırdan eğitildi "
        f"fark etmez — bu 3 dosya HER çalıştırmada yeniden üretilir)\n"
        f"   Tahmin sayısı  : {len(df_ortools)}\n"
        f"   Tarih aralığı  : {PREDICT_START} → {PREDICT_END}\n"
        f"   Toplam Talep Tahmini (desi) : {total_desi_tahmini:,.3f}\n"
        f"   1) {OUTPUT_CSV}\n"
        f"   2) {OUTPUT_EXCEL}\n"
        f"   3) {OUTPUT_TALEP_XLSX}\n"
        f"============================================================"
    )
    print("\n📋 OR-Tools Payload Örnek (İlk 5 Satır):")
    print(df_ortools.head().to_string(index=False))
    print(f"\n🔢 Toplam Talep Tahmini (tüm rota/tarih/saat dilimi): {total_desi_tahmini:,.3f} desi")

    return df_ortools


# ---------------------------------------------------------------------------
# Örnek payload çıktısı
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # save_json=True: main.py -> forecast_payload.run_predict_model() bu betiği
    # subprocess olarak çalıştırıp alns_payload.json'u okuyor. save_json=False
    # kalırsa JSON hiç yenilenmez ve (slot bilgisi olmayan, eski tarihli) bayat
    # bir dosya okunmaya devam eder — Faz 2 slot bazlı optimizasyonun önkoşulu.
    run(save_json=True)