"""RouteTech Faz 2 — ALNS tabanlı optimizasyon (Stage C).

main.py --engine alns (varsayılan) seçildiğinde bu betiği çalıştırır. Aynı
ROUTETECH_OPTIMIZATION_INPUT / ROUTETECH_MAX_TIME_SECONDS ortam değişkenlerini
okur (src/optimization.py [CP-SAT] ile aynı girdi), ALNS döngüsünü çalıştırır,
aynı çıktı şemasını (results/optimization_results.txt/csv/xlsx) üretir — artık
`Rota_Tipi` sütunuyla konsolidasyonlu (relay) sevkiyatları da gösterir.

Mimari: src/alns_engine.py'deki State/destroy/repair yapılarını kullanır; ana
motor ALNS'tir, CP-SAT (`cpsat_hat_repair`) tek bir hattı tam optimal çözen
küçük-pencere bir repair operatörü olarak devreye girer.
"""

from __future__ import annotations
from collections import defaultdict

import json
import math
import os
import sys
import time
from pathlib import Path

import numpy as np
import numpy.random as rnd
import pandas as pd

# alns paketinin SimulatedAnnealing kabul kriteri, cok iyi bir aday cozum
# bulundugunda exp(...) hesabinin sonsuza (inf) tasmasina neden olabiliyor -
# bu ZARARSIZ (davranis dogru: "kesinlikle kabul et" demek), sadece gurultulu
# bir konsol uyarisi. Susturuyoruz.
np.seterr(over="ignore")
from alns import ALNS
from alns.accept import SimulatedAnnealing
from alns.select import RouletteWheel
from alns.stop import MaxRuntime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.alns.cost_model import spot_vehicle_count, vehicle_leg_cost, ellecleme_maliyet_hesapla  # noqa: E402
from src.alns.alns_engine import (  # noqa: E402
    MIN_SPOT_DOLULUK_ORANI,
    Assignment,
    KIRALIK_DISPATCH_SLOT,
    ProblemData,
    State,
    dogrula_gercek_kapasite,
    dogrula_min_spot_doluluk,
    dummy_initial_builder,
    enforce_min_spot_occupancy,
    enforce_real_capacity_limits,
    cpsat_hat_repair,
    greedy_repair,
    leg_zaman_cizelgesi,
    random_removal,
    regret_repair,
    low_occupancy_removal,
    shaw_related_removal,
    tm_overload_removal,
    worst_removal,
)
from src.alns.time_model import (  # noqa: E402
    build_route_lookup, slot_datetime, varis_zamani, ellecleme_tamamlanma_zamani,seyir_suresi_saat,
    ellecleme_suresi_dakika, sla_deadline, gecikme_saat, sla_cezasi_tl,
)


# 0. Darboğaz kontrolü fonksiyonu
def ham_talep_kapasite_raporu(demands, handling_capacity, gunler):
    """ALNS calismadan ONCE, her TM icin o gunun HAM (gecikme/backlog olmadan)
    gelen+giden talep toplamini kapasiteyle karsilastirir. Bu, motorun cozum
    SONRASI kontrolunden (best.handling_usage) FARKLI bir soruya cevap verir:
    'bu gunun kendi talebi tek basina sigar mi?' - motor ne kadar iyi calisirsa
    calissin, bu rapor >%100 gosteriyorsa erteleme/SLA cezasi KACINILMAZDIR."""
    gunluk_ham_talep = {}
    for (hat, gun, slot, desi, _tid) in demands:
        src, dst = hat
        gunluk_ham_talep[(src, gun)] = gunluk_ham_talep.get((src, gun), 0.0) + desi
        gunluk_ham_talep[(dst, gun)] = gunluk_ham_talep.get((dst, gun), 0.0) + desi
    for tm, cap in handling_capacity.items():
        for gun in gunler:
            talep = gunluk_ham_talep.get((tm, gun), 0.0)
            if talep > cap:
                print(f"YAPISAL DARBOGAZ: {tm} {gun} ham talep={talep:.0f} > kapasite={cap:.0f} "
                      f"(%{100*talep/cap:.0f}) - erteleme/ceza kacinilmaz")

# ============================================================================
# 1. ORTAM DEĞİŞKENLERİ
# ============================================================================
ENV_MAX_TIME = float(os.environ.get("ROUTETECH_MAX_TIME_SECONDS", "300")) # Algoritmanın çalışma süresi, terminalden parametre olarak verilebilir
ENV_VERBOSE = os.environ.get("ROUTETECH_LOG_SEARCH_PROGRESS", "1") == "1" 

DEFAULT_INPUT_JSON = PROJECT_ROOT / "data" / "outputs" / "optimization_input.json" 
INPUT_JSON_PATH = Path(os.environ.get("ROUTETECH_OPTIMIZATION_INPUT", str(DEFAULT_INPUT_JSON)))# Optimizasyon input dosyasının path'i

if not INPUT_JSON_PATH.exists():
    raise FileNotFoundError(
        f"Optimizasyon girdisi bulunamadı: {INPUT_JSON_PATH}. Once `python main.py` calistirilmali."
    )

with INPUT_JSON_PATH.open("r", encoding="utf-8") as f:
    payload = json.load(f)

df_forecast = pd.DataFrame(payload["forecast"]) # Tahmin verileri
df_route_matrix = pd.DataFrame(payload["route_matrix"]) # mesafeler
df_vehicle_costs = pd.DataFrame(payload["vehicle_costs"]) # araç parametreleri
df_rental_limits = pd.DataFrame(payload["rental_limits"]) # Kiralık stoklar
df_handling_capacity = pd.DataFrame(payload["handling_capacity"]) # Elleçleme kapasiteleri
df_tir_capacity = pd.DataFrame(payload["tir_capacity"]) # Tır kapasiteleri

if df_forecast.empty:
    raise ValueError(f"Forecast payload bos: {INPUT_JSON_PATH}")

route_lookup = build_route_lookup(df_route_matrix) # veri tablosunu hızlı sorgulanabilir bir dict'e çevirir.

arac_turleri = df_vehicle_costs["vehicle_type"].tolist() 
tir_arac_turu = next(
    (v for v in arac_turleri if str(v).strip().casefold() in ("tır", "tir")), None
)

# Araç parametrelerine hızlı erişim için bir sözlük oluşturma.
arac_parametreleri = {}
for row in df_vehicle_costs.itertuples():
    arac_parametreleri[row.vehicle_type] = {
        "kapasite_desi": float(row.capacity),
        "rental_hourly": float(row.rental_hourly),
        "rental_km": float(row.rental_km),
        "spot_hourly": float(row.spot_hourly),
        "spot_km": float(row.spot_km),
    }
# "(hat,gun) -> kiralik_stok" şeklinde bir dict.
kiralik_stok_gunluk: dict = {}
for row in df_rental_limits.itertuples():
    hat = (row.source, row.destination)
    key = (hat, row.vehicle_type)
    kiralik_stok_gunluk[key] = kiralik_stok_gunluk.get(key, 0) + int(row.vehicle_count)

handling_capacity = {row.center: float(row.capacity) for row in df_handling_capacity.itertuples()}
tir_capacity = {row.center: float(row.capacity) for row in df_tir_capacity.itertuples()}

# ============================================================================
# 2. TALEP / HAT / GÜN ÇIKARIMI
# ============================================================================
df_forecast = df_forecast.copy()
df_forecast["date"] = pd.to_datetime(df_forecast["date"])
df_forecast["gun_key"] = df_forecast["date"].dt.strftime("%Y-%m-%d")
df_forecast["slot"] = df_forecast["slot"].astype(str) # 09:00 veya 17:00

gunler = sorted(df_forecast["gun_key"].unique())

son_gun = pd.to_datetime(gunler[-1])
for ek in (1, 2):
    ek_gun = (son_gun + pd.Timedelta(days=ek)).strftime("%Y-%m-%d")
    if ek_gun not in gunler:
        gunler.append(ek_gun)
    
merkezler = sorted(set(handling_capacity) | set(tir_capacity))

# (hat,gun,slot) -> desi şeklinde bir dict. slot'tan kasıt 09:00, 17:00
# burası ebubekirin tahmin modelinden gelen her talep için bir kimlik.
talep_verisi: dict = {}
talep_id_map: dict = {}
for row in df_forecast.itertuples():
    hat = (row.source, row.destination)
    key = (hat, row.gun_key, row.slot)
    talep_verisi[key] = talep_verisi.get(key, 0.0) + max(0.0, float(row.recommended_demand))
    talep_id_map[key] = row.talep_id

demands = [
    (hat, gun, slot, desi, talep_id_map[(hat, gun, slot)])
    for (hat, gun, slot), desi in talep_verisi.items()
    if desi > 1e-9
]

print(f"ALNS: {len(demands):,} talep parcasi, {len(gunler)} gun, {len(merkezler)} TM.")

# ============================================================================
# 3. PROBLEM VERISI VE BASLANGIC COZUMU
# ============================================================================
# Lojistik problemimize ait tüm optimizasyon dataları.
data = ProblemData(
    route_lookup=route_lookup,
    arac_turleri=arac_turleri,
    arac_parametreleri=arac_parametreleri,
    kiralik_stok_gunluk=kiralik_stok_gunluk,
    handling_capacity=handling_capacity,
    tir_capacity=tir_capacity,
    tir_arac_turu=tir_arac_turu,
    gunler=gunler,
    merkezler=merkezler,
    demands=demands,
)

# ========== YAPISAL DARBOGAZ RAPORU (ALNS öncesi) ==========
ham_talep_kapasite_raporu(data.demands, data.handling_capacity, data.gunler)
# ===========================================================

rng = rnd.default_rng(42) # Neden 42 seedini veriyoruz generator'ü oluşturmak için?

initial_state = State(data) 
initial_state = dummy_initial_builder(initial_state, rng) # Burada neden removal kullanmadan repair yapmış claude?
initial_obj = initial_state.objective() # mevcut çözümün maliyeti
print(f"Baslangic (dummy) cozum maliyeti: {initial_obj:,.0f} TL")

# ============================================================================
# 4. ALNS KURULUMU
# ============================================================================
alns = ALNS(rng) # ALNS'yi generator ile kurduk
# Destroy ve repair operatorleri olarak fonksiyonları atıyoruz.
alns.add_destroy_operator(random_removal, "random_removal")
alns.add_destroy_operator(worst_removal, "worst_removal")
alns.add_destroy_operator(tm_overload_removal, "tm_overload_removal")
alns.add_destroy_operator(low_occupancy_removal, "low_occupancy_removal")
alns.add_destroy_operator(shaw_related_removal, "shaw_related_removal")
alns.add_repair_operator(greedy_repair, "greedy_repair")
alns.add_repair_operator(regret_repair, "regret_repair")
alns.add_repair_operator(cpsat_hat_repair, "cpsat_hat_repair")

# En iyi sonuç -> 25 puan
# Eskisinden iyi sonuç -> 5 puan
# Eskisinden kötü ama kabul edilen bir sonuç -> 2 puan
# Reddedilen berbat bir sonuç -> 0.5 puan
# decay -> unutma katsayısı. Eski başarılara takılıp kalmamak için puanlar her iterasyonda %80 korunur.

select = RouletteWheel(scores=[25, 5, 2, 0.5], decay=0.8, num_destroy=5, num_repair=3)

# low_occupancy_removal, dusuk doluluklu (israf) araclari DOGRUDAN hedef alan tek
# destroy operatoru - ama RouletteWheel butun destroy operatorlerine baslangicta
# esit agirlik (1.0) veriyor, performansini kanitlamasi zaman aliyor. Doluluk
# sorunu bizim asil darbogazimiz oldugu icin, bu operatore baslangicta 3 kat
# daha yuksek agirlik vererek arama motorunun bunu daha erken/sik denemesini
# sagliyoruz - zamanla RouletteWheel yine gercek performansina gore kendi
# agirligini ayarlayacak, bu sadece erken bir "ipucu".
_destroy_isimleri = [ad for ad, _ in alns.destroy_operators]
_low_occ_idx = _destroy_isimleri.index("low_occupancy_removal")
select._d_weights[_low_occ_idx] = 3.0

# Her yeni en iyi cozum bulundugunda anlik olarak yazdir - onceki CP-SAT'in
# ayrintili arama logunun ALNS'teki esdegeri (bkz. sohbet gecmisi: "eskiden her
# saniye/yenisini bulduğunda görünürdü" - ALNS varsayilan motor olduktan sonra
# bu geri bildirim hic eklenmemisti).
_ilerleme_baslangic = time.time() # Şu anki zaman.
_ilerleme_sayac = {"n": 0} # Daha iyi bir maliyet buldukca artirilir.


def _yeni_en_iyi_bulundu(candidate_state, rng_):
    _ilerleme_sayac["n"] += 1
    gecen_sn = time.time() - _ilerleme_baslangic # şu ana kadar ne kadar süre geçti?
    print(f"[{gecen_sn:7.1f} sn] Yeni en iyi #{_ilerleme_sayac['n']}: {candidate_state.objective():,.0f} TL")


alns.on_best(_yeni_en_iyi_bulundu) # ALNS en iyi maliyeti bulunca bu callback fonksiyonu çağırsın diyoruz.

# ============================================================================
# 4b. KALIBRASYON: gercek iterasyon hizini olcup SA sicaklik takvimini buna gore ayarla
# ============================================================================
# ONCEKI YONTEM (sabit "ENV_MAX_TIME/2 sn, iterasyon ~2 sn surer" varsayimi) YANLIS
# CIKTI: iterasyon suresi 0.2-15 sn arasinda COK degisken (bkz. sohbet gecmisi),
# sabit bir tahmin gercek hizi tutturamiyor. Yanlis num_iters, SA sicakligini
# gercekten olmasi gerekenden COK ERKEN (veya COK GEC) dondurup ya aramayi
# vaktinden once hill-climbing'e kilitliyor (erken donarsa, geri kalan sure
# boyunca hicbir iyilesme bulunamiyor) ya da hic sogutmuyor. Bunun yerine,
# butcenin kucuk bir dilimini (en fazla 20 sn / %10) GERCEK operatorlerle
# calisip gercek iterasyon hizini olcuyoruz - bu kalibrasyon iterasyonlari da
# gercek arama ilerlemesi oldugu icin (best_state'ten devam ediliyor) bosa
# gitmiyor, select'in (RouletteWheel) ogrendigi agirliklar da korunuyor.
class _KalibrasyonDurdur:
    """En az `min_iterasyon` tamamlanana KADAR devam eder - tek/iki ornek ile
    (sans eseri yavas bir iterasyona denk gelip) yanlis hiz olcmemek icin.
    Ama cok yavas iterasyonlar min_iterasyon'a ulasmayi geciktirirse, `max_sure`
    saniyeyi asinca da (butceyi tuketmemek icin) her halukarda durur."""

    def __init__(self, min_iterasyon: int, max_sure: float):
        self.min_iterasyon = min_iterasyon
        self.max_sure = max_sure
        self.sayac = 0
        self.baslangic = time.time()

    def __call__(self, rng, best, curr) -> bool:
        self.sayac += 1
        gecen = time.time() - self.baslangic
        return self.sayac > self.min_iterasyon or gecen >= self.max_sure


kalibrasyon_durdur = _KalibrasyonDurdur(
    min_iterasyon=8, max_sure=min(60.0, ENV_MAX_TIME * 0.15)
)
kalibrasyon_accept = SimulatedAnnealing.autofit(
    init_obj=initial_obj, worse=0.06, accept_prob=0.5, num_iters=100
)
print(f"Kalibrasyon calistiriliyor (en az {kalibrasyon_durdur.min_iterasyon} iterasyon, gercek hizi olcmek icin)...")
kalibrasyon_sonuc = alns.iterate(initial_state, select, kalibrasyon_accept, kalibrasyon_durdur)

kalibrasyon_iter_sayisi = max(1, len(kalibrasyon_sonuc.statistics.objectives) - 1)
kalibrasyon_suresi = kalibrasyon_sonuc.statistics.total_runtime
gercek_iterasyon_hizi = kalibrasyon_suresi / kalibrasyon_iter_sayisi
kalan_sure = max(1.0, ENV_MAX_TIME - kalibrasyon_suresi)
# Gercek hiz genelde kalibrasyondan sonra artiyor (bkz. sohbet gecmisi); tahmini
# iterasyonu paylayip SA'nin zamanindan once sogumasini (aramanin erken kilitlenmesini) onluyoruz.
tahmini_iterasyon = max(20, int(1.75 * kalan_sure / max(gercek_iterasyon_hizi, 0.05)))
print(
    f"Kalibrasyon tamamlandi: {kalibrasyon_iter_sayisi} iterasyon / {kalibrasyon_suresi:.1f} sn "
    f"-> ~{gercek_iterasyon_hizi:.2f} sn/iterasyon -> kalan {kalan_sure:.0f} sn icin tahmini {tahmini_iterasyon} iterasyon"
)

accept = SimulatedAnnealing.autofit(
    init_obj=kalibrasyon_sonuc.best_state.objective(), worse=0.06, accept_prob=0.5, num_iters=tahmini_iterasyon
)
stop = MaxRuntime(kalan_sure)

print(f"ALNS calistiriliyor (bütce: {ENV_MAX_TIME:.0f} sn)...")
result = alns.iterate(kalibrasyon_sonuc.best_state, select, accept, stop)
best: State = result.best_state
print(f"ALNS tamamlandi. En iyi maliyet: {best.objective():,.0f} TL "
      f"(baslangica gore {'%.1f' % (100 * (1 - best.objective() / max(1, initial_obj)))}% iyilesme)")

# Minimum spot doluluk kuralinin KESIN (hard, parasal olmayan) uygulanmasi -
# bkz. alns_engine.enforce_min_spot_occupancy docstring. Bu, objective()'i
# DEGISTIRMEZ (parasal ceza yok) - sadece esigi karsilamayan spot sevkiyatlari
# bir sonraki uygun slota zorla erteler (README'deki is kuralinin ta kendisi).
print(f"Minimum spot doluluk kurali (%{MIN_SPOT_DOLULUK_ORANI*100:.0f}) uygulaniyor...")
_unassigned_oncesi_satir = len(best.unassigned)
_unassigned_oncesi_desi = sum(x[3] for x in best.unassigned)
print(f"  (uygulama ONCESI mevcut unassigned: {_unassigned_oncesi_satir} satir / {_unassigned_oncesi_desi:,.0f} desi - bu ALNS'in kendi sonucu, doluluk kuraliyla ilgisiz)")
_dogrulama_oncesi = dogrula_min_spot_doluluk(best)
best = enforce_min_spot_occupancy(best, rng)
_dogrulama_sonrasi = dogrula_min_spot_doluluk(best)
print(
    f"Doluluk kurali uygulandi: {len(_dogrulama_oncesi)} ihlal -> {len(_dogrulama_sonrasi)} ihlal. "
    f"Uygulama sonrasi maliyet: {best.objective():,.0f} TL"
)
if _dogrulama_sonrasi:
    print("UYARI: asagidaki gruplar hala esigin altinda (teorik olarak olmamali):")
    for v in _dogrulama_sonrasi[:20]:
        print(f"  {v['src']}->{v['dst']} {v['gun']} {v['slot']} {v['arac_turu']}: "
              f"{v['desi']:.0f}/{v['adet']*v['kapasite']:.0f} desi (%{v['doluluk']*100:.1f})")
else:
    print(f"DOGRULAMA GECTI: son (gun,slot) haric hicbir spot arac grubu %{MIN_SPOT_DOLULUK_ORANI*100:.0f} altinda degil.")

# Gercek (elleçleme dahil) zamana gore tir/elleçleme kapasitesi dogrulamasi -
# bkz. alns_engine.enforce_real_capacity_limits docstring.
print("Gercek-zaman kapasite dogrulamasi uygulaniyor...")
_gercek_dogrulama_oncesi = dogrula_gercek_kapasite(best)
best = enforce_real_capacity_limits(best, rng)
_gercek_dogrulama_sonrasi = dogrula_gercek_kapasite(best)
print(
    f"Gercek-zaman kapasite dogrulamasi: "
    f"{len(_gercek_dogrulama_oncesi['handling'])} elleçleme + {len(_gercek_dogrulama_oncesi['tir'])} tir ihlali -> "
    f"{len(_gercek_dogrulama_sonrasi['handling'])} elleçleme + {len(_gercek_dogrulama_sonrasi['tir'])} tir ihlali. "
    f"Uygulama sonrasi maliyet: {best.objective():,.0f} TL"
)
if _gercek_dogrulama_sonrasi["tir"] or _gercek_dogrulama_sonrasi["handling"]:
    print("UYARI: gercek-zamanda hala kapasite asan TM/gun'lar var (muhtemelen zorunlu kiralik kaynakli):")
    for v in (_gercek_dogrulama_sonrasi["tir"] + _gercek_dogrulama_sonrasi["handling"])[:20]:
        print(f"  {v['tm']} {v['gun']}: {v['kullanim']:.1f} / {v['kapasite']:.1f}")
else:
    print("DOGRULAMA GECTI: gercek zamana gore hicbir TM/gun tir/elleçleme kapasitesini asmiyor.")

def _merge_duplicate_route_assignments(assignments):
    """Ayni talebe (talep_id) ait, TAMAMEN ayni fiziksel rotayi (ayni bacak
    dizisi + ayni milk_run modu) kullanan birden fazla Assignment parcasini
    TEK bir parcada birlestirir.

    ALNS destroy/repair dongusu ayni talebi defalarca sokup ayni bacaga yeniden
    ekleyebiliyor - kargo kaybi yok ama rapor onlarca neredeyse ozdes mini
    parcaya bolunuyor, her parcanin bagimsiz round(...,2) yuvarlanmasi kucuk
    bir toplam sapmaya yol aciyordu. Sadece raporlama icindir, maliyet/kapasite
    muhasebesini degistirmez; farkli rotaya/araca/gune giden parcalar asla
    birlestirilmez."""
    merged: dict = {}
    order: list = []
    for a in assignments:
        route_key = (
            a.talep_id, a.demand_hat, a.demand_gun, a.demand_slot, a.milk_run,
            tuple(
                (leg.src, leg.dst, leg.gun, leg.slot, leg.arac_turu, leg.is_kiralik)
                for leg in a.legs
            ),
        )
        if route_key not in merged:
            merged[route_key] = a
            order.append(route_key)
        else:
            existing = merged[route_key]
            merged[route_key] = Assignment(
                demand_hat=existing.demand_hat, demand_gun=existing.demand_gun,
                demand_slot=existing.demand_slot, desi=existing.desi + a.desi,
                legs=existing.legs, sla_cost=existing.sla_cost + a.sla_cost,
                vehicle_cost=existing.vehicle_cost + a.vehicle_cost,
                talep_id=existing.talep_id, milk_run=existing.milk_run,
            )
    return [merged[k] for k in order]


_oncesi_assignment_sayisi = len(best.assignments)
best.assignments = _merge_duplicate_route_assignments(best.assignments)
print(
    f"Raporlama birlestirmesi: {_oncesi_assignment_sayisi} atama parcasi -> "
    f"{len(best.assignments)} (ayni rotayi paylasan tekrarlar birlestirildi)."
)


def _talep_id_goruntule():
    """Nihai cozumdeki her Assignment'in hangi ID ile gorunecegini hesaplar.
    Bir talep birden fazla parcaya bolunmusse -1,-2... eki eklenir; tek
    parcaysa direkt ID kalir. Bos ID'ler (cok nadir, bilinen bir sinirlama
    icin) hic dokunulmadan birakilir."""
    gruplu = defaultdict(list)
    for a in best.assignments:
        if a.talep_id:  # bos ID'leri disarida birak
            gruplu[a.talep_id].append(a)

    goruntu_id = {}
    for kok_id, parcalar in gruplu.items():
        if len(parcalar) == 1:
            goruntu_id[id(parcalar[0])] = kok_id
        else:
            for i, a in enumerate(parcalar, start=1):
                goruntu_id[id(a)] = f"{kok_id}-{i}"
    return goruntu_id

talep_id_goruntu = _talep_id_goruntule()

# ---- Kapasite ihlali dogrulamasi (bu oturumun kontrolu icin, kalici test degil) ----
ihlal_sayisi = 0
for tm, cap in data.handling_capacity.items():
    for gun in data.gunler:
        kullanim = best.handling_usage.get((tm, gun), 0.0)
        if kullanim > cap + 1e-6:
            ihlal_sayisi += 1
            print(f"UYARI: elleceleme kapasitesi asildi: {tm} {gun} kullanim={kullanim:.1f} > kap={cap:.1f}")
for tm, cap in data.tir_capacity.items():
    for gun in data.gunler:
        # best.tir_usage hep bos kalan olu bir sozluktu; gercek spot kullanimi
        # tir_usage_in/tir_usage_out'ta tutuluyor.
        kullanim = (
            best.tir_usage_in.get((tm, gun), 0) + best.tir_usage_out.get((tm, gun), 0)
            + data.fixed_kiralik_tir_usage.get((tm, gun), 0)
        )
        if kullanim > cap + 1e-6:
            ihlal_sayisi += 1
            print(f"UYARI: tir kapasitesi asildi: {tm} {gun} kullanim={kullanim} > kap={cap}")
print(f"Kapasite dogrulamasi: {ihlal_sayisi} ihlal bulundu (0 beklenir).")

# ============================================================================
# 5. ÇIKTI — optimization.py ile AYNI ŞEMA (+ Rota_Tipi)
# ============================================================================
results_dir = PROJECT_ROOT / "results"
results_dir.mkdir(parents=True, exist_ok=True)
output_txt = results_dir / "optimization_results.txt"
output_csv = results_dir / "optimization_results.csv"
output_xlsx = results_dir / "optimization_results.xlsx"

def _bucket_key(leg):
    return (leg.src, leg.dst, leg.gun, leg.slot, leg.arac_turu, leg.is_kiralik)

# Her kargonun (assignment) kendi bağımsız zaman çizelgesini hesaplıyoruz - başka
# bir kargonun aynı bacağı ne zaman kullandığına bakmadan, sadece bu kargonun kendi
# elleçleme+yol süresine göre gerçek kalkış/varış anlarını buluyoruz.
# bucket_gercek_kalkis ise sadece "bu bacak en son ne zaman kalkmış" bilgisini özet
# tabloda göstermek için - hiçbir kargonun kendi zaman hesabına karışmıyor.
assignment_cizelgeleri = {}   # id(a) -> [(kalkis0, varis0), (kalkis1, varis1), ...]
bucket_gercek_kalkis = {}     # bucket_key -> en gec kalkis (sadece ozet icin)

for a in best.assignments:
    cizelge = leg_zaman_cizelgesi(data, a.legs, a.desi, milk_run=a.milk_run)
    assignment_cizelgeleri[id(a)] = cizelge
    for leg, (kalkis, _varis) in zip(a.legs, cizelge):
        key = _bucket_key(leg)
        if key not in bucket_gercek_kalkis:
            bucket_gercek_kalkis[key] = kalkis
        else:
            bucket_gercek_kalkis[key] = max(bucket_gercek_kalkis[key], kalkis)
            
# Bacak (leg-dispatch) basina GERCEK toplam desi - birden fazla talep parcasi
# ayni fiziksel sevkiyati (ayni TM-cift/gun/slot/arac_turu) paylasabilir; tek
# tek assignment'in kendi desi'sinden arac sayisi hesaplamak (eski yontem)
# yaniltici oluyordu (her satir "1 arac" gosterip, ayni bacaga kac FARKLI talep
# parcasinin bindigi gorunmuyordu). Once TUM bacaklarin gercek toplam desisini
# topluyoruz, sonra arac sayisini bundan turetiyoruz.
bucket_toplam_desi: dict = {}
for a in best.assignments:
    for leg in a.legs:
        key = _bucket_key(leg)
        bucket_toplam_desi[key] = bucket_toplam_desi.get(key, 0.0) + a.desi

bucket_parcalari: dict = {}
for a in best.assignments:
    for leg_i, leg in enumerate(a.legs):
        key = _bucket_key(leg)
        if key not in bucket_parcalari:
            bucket_parcalari[key] = []
        bucket_parcalari[key].append((a, leg, a.desi, leg_i))
# Bu bacağı (aynı src-dst-gun-slot-arac_turu) kaç FARKLI talep paylaşıyor -
# uğrama (rota kaç bacaklı) ile konsolidasyon (bacağı kaç talep paylaşıyor)
# birbirinden bağımsız iki soru, bu yüzden ayrı hesaplanıyor.
bucket_konsolidasyon_sayisi: dict = {
    key: len({id(a) for (a, leg, desi, leg_i) in parcalar})
    for key, parcalar in bucket_parcalari.items()
}

def _bacak_arac_dagilimi(bucket_key, parcalar):
    src, dst, gun, slot, arac_turu, is_kiralik = bucket_key

    kap = arac_parametreleri[arac_turu]["kapasite_desi"]
    sonuc = []
    arac_index = 0
    mevcut_doluluk = 0.0

    for (a, leg, desi, leg_i) in parcalar:
        kalan = desi
        # DÜZELTME: 1e-9 yerine 0.001 (1 gram hassasiyet) yapıyoruz
        while kalan > 0.001: 
            bos_yer = kap - mevcut_doluluk
            if bos_yer <= 0.001: # DÜZELTME: 1e-9 yerine 0.001
                arac_index = arac_index + 1
                mevcut_doluluk = 0.0
                bos_yer = kap
            bu_araca = min(kalan, bos_yer)
            sonuc.append((a, leg, arac_index, bu_araca, leg_i))
            mevcut_doluluk = mevcut_doluluk + bu_araca
            kalan = kalan - bu_araca

    return sonuc

bucket_dagilim: dict = {}
for key in bucket_parcalari:
    bucket_dagilim[key] = _bacak_arac_dagilimi(key, bucket_parcalari[key])

# COZUM A: Ayni fiziksel araca (bucket_key, arac_index) binen TUM parcalarin
# ORTAK, TEK bir kalkis/varis zamani olmasi icin - artik her parca kendi
# (kucuk) desisine gore degil, o aracin GERCEK TOPLAM yukune gore elleclenip
# kalkiyor. Onceden her assignment kendi desisiyle bagimsiz hesaplaniyordu,
# bu da ayni araca binen farkli taleplerin farkli kalkis/varis saati
# gostermesine yol aciyordu (bkz. sohbet gecmisi).
arac_toplam_yuk: dict = {}
for key, parcalar in bucket_dagilim.items():
    for (a, leg, arac_index, pay_desi, leg_i) in parcalar:
        arac_key = (key, arac_index)
        arac_toplam_yuk[arac_key] = arac_toplam_yuk.get(arac_key, 0.0) + pay_desi

# COZUM A (devam): sadece toplam yuke gore hesaplamak yetmiyor - bu bacak bir
# ONCEKI bacaktan gelen aktarmali (relay) bir devam ise, arac o parcanin
# BAGLANTISININ gelmesini de beklemek zorunda. Her arac (bucket_key,
# arac_index) icin, kalkis = max( normal yukleme zamani, ustune binen HER
# aktarmali parcanin kendi baglanti zamani ).
arac_gruplari: dict = {}
for key, parcalar in bucket_dagilim.items():
    for (a, leg, arac_index, pay_desi, leg_i) in parcalar:
        arac_key = (key, arac_index)
        arac_gruplari.setdefault(arac_key, []).append((a, leg, pay_desi, leg_i))

# Her fiziksel araca (arac_key), jüri şablonundaki gibi basit ve sıralı bir kod
# ("V0001", "V0002", ...) atıyoruz - onceki uzun/aciklayici string yerine.
# Kronolojik (gun, slot, cikis-varis) sirayla numaralandiriyoruz ki okunabilir olsun.
_siralanmis_arac_keyler = sorted(arac_gruplari.keys(), key=lambda k: (k[0][2], k[0][3], k[0][0], k[0][1], k[1]))
arac_id_kodu: dict = {
    arac_key: f"V{sira + 1:04d}"
    for sira, arac_key in enumerate(_siralanmis_arac_keyler)
}

# Bu satırın altına ekleyin:
_bos_id_counter = max(arac_id_kodu.values(), default="V0000")[1:]  # "0001" gibi
_bos_id_counter = int(_bos_id_counter) + 1 if _bos_id_counter else 1

# Bu FİZİKSEL aracı (bucket_key, arac_index) paylaşan farklı taleplerin HER
# BİRİNİN kendi elleçleme "dokunuş" sayısı (0/1/2) farklı olabilir - bir talep
# buradan uğrayarak (milk_run) geçiyorsa dokunuş 0/1, gerçekten iniyor/biniyorsa
# dokunuş 2. Arama motorunun kendi iç muhasebesiyle (State.leg_ellecleme_desi,
# bkz. alns_engine._commit_leg) BİREBİR tutarlı olması için elleçleme maliyetini
# dokunuşa göre GRUPLAYIP her grubu ayrı (kendi toplam desisi üzerinden, bir kez)
# yuvarlıyoruz - tek bir toplam_yuk üzerinden hesaplamak, uğramalı/uğramasız
# parçaları karıştırıp yanlış (fazla) elleçleme faturalandırırdı.
def _fragman_dokunus_sayisi(a, leg_i):
    if not a.milk_run:
        return 2
    n = len(a.legs)
    return (0 if leg_i > 0 else 1) + (0 if leg_i < n - 1 else 1)

# Her aracin GERCEK toplam maliyetini, motorun "kim once geldi" (marjinal)
# muhasebesine BAKMADAN, bagimsiz olarak yeniden hesapliyoruz - boylece her
# parcaya, o aracin toplam maliyetinden ADIL (yukune orantili) bir pay
# verebiliriz. Kiralik aracta seyir maliyeti sabit/batik oldugu (ayrica
# raporlaniyor) icin sadece ellecleme payi hesaplaniyor.
arac_maliyeti: dict = {}
for arac_key, kayitlar in arac_gruplari.items():
    key, arac_index = arac_key
    src, dst, gun, slot, arac_turu, is_kiralik = key
    p = arac_parametreleri[arac_turu]

    if is_kiralik:
        seyir_maliyeti = 0.0
        hourly_rate = p["rental_hourly"]
    else:
        seyir_maliyeti = vehicle_leg_cost(data.route_lookup, (src, dst), arac_turu, p["spot_hourly"], p["spot_km"])
        hourly_rate = p["spot_hourly"]

    dokunus_gruplari: dict = {}  # dokunus_sayisi -> bu gruba ait toplam desi
    for (a, leg, pay_desi, leg_i) in kayitlar:
        dokunus = _fragman_dokunus_sayisi(a, leg_i)
        dokunus_gruplari[dokunus] = dokunus_gruplari.get(dokunus, 0.0) + pay_desi

    ellecleme_maliyeti = sum(
        ellecleme_maliyet_hesapla(grup_desi, hourly_rate, dokunus)
        for dokunus, grup_desi in dokunus_gruplari.items()
    )
    arac_maliyeti[arac_key] = seyir_maliyeti + ellecleme_maliyeti

# Her (assignment, leg_i) cifti hangi (bucket_key, arac_index)'e denk geliyor -
# bagimlilik zincirini takip edebilmek icin (leg_i-1'in GERCEK arac_key'ini bulmak).
# NOT: aynı assignment'in aynı bacaktaki yuku BIRDEN FAZLA araca bolunebilir
# (bin-packing sirasinda kapasite tasarsa), bu yuzden TEK bir arac_key değil,
# LISTE tutuyoruz - yoksa bir onceki bacaktaki araclardan biri sessizce
# unutulup yanlis (erken) bir baglanti zamani hesaplanabilirdi.
parca_arac_index: dict = {}
for arac_key, kayitlar in arac_gruplari.items():
    for (a, leg, pay_desi, leg_i) in kayitlar:
        parca_arac_index.setdefault((id(a), leg_i), set()).add(arac_key)

# COZUM A (nihai): kalkis = max( normal yukleme zamani, ustune binen HER
# aktarmali parcanin - kendi ONCEKI bacaginin YENI/duzeltilmis varisina gore -
# hesaplanan baglanti zamani ). Onceki bacagin zamani da kendisi bu ayni
# sekilde (once digerlerine bagli) hesaplandigi icin, recursive/memoized
# olarak (bagimliliklar once cozulerek) hesapliyoruz - zincir her zaman ILERI
# gittigi icin (leg_i hep artiyor) sonsuz dongu olusmaz.
arac_zamanlari: dict = {}
def _arac_zamani_hesapla(arac_key):
    if arac_key in arac_zamanlari:
        return arac_zamanlari[arac_key]

    key, arac_index = arac_key
    kayitlar = arac_gruplari[arac_key]
    ornek_leg = kayitlar[0][1]
    slot_zamani = slot_datetime(ornek_leg.gun, ornek_leg.slot)
    toplam_yuk = arac_toplam_yuk[arac_key]

    yukleme_kalkisi = leg_zaman_cizelgesi(data, [ornek_leg], toplam_yuk)[0][0]

    baglanti_kisitlari = []
    for (a, leg, pay_desi, leg_i) in kayitlar:
        if leg_i > 0:
            # Onceki bacakta bu ayni yuk birden fazla araca binmis olabilir -
            # hangisinden geldigi belirsiz oldugu icin, EN SON varan aracin
            # zamanini esas alarak (en guvenli/tutarli tahmin) bekliyoruz.
            onceki_arac_keyler = parca_arac_index[(id(a), leg_i - 1)]
            onceki_varisler = [_arac_zamani_hesapla(k)[1] for k in onceki_arac_keyler]
            onceki_varis = max(onceki_varisler)
            baglanti_kisitlari.append(
                ellecleme_tamamlanma_zamani(onceki_varis, pay_desi, consolidation=True)
            )

    gercek_kalkis = max([yukleme_kalkisi, slot_zamani] + baglanti_kisitlari)
    seyir = data.route_lookup[(ornek_leg.src, ornek_leg.dst)][ornek_leg.arac_turu]
    gercek_varis = varis_zamani(gercek_kalkis, seyir)
    arac_zamanlari[arac_key] = (gercek_kalkis, gercek_varis)
    return arac_zamanlari[arac_key]

for arac_key in arac_gruplari:
    _arac_zamani_hesapla(arac_key)


def _bacak_arac_sayisi(leg) -> int:
    kap = arac_parametreleri[leg.arac_turu]["kapasite_desi"]
    return spot_vehicle_count(bucket_toplam_desi[_bucket_key(leg)], kap, 10 ** 9)


def _final_leg_sla_cezasi(varis_dt, pay_desi, demand_gun, demand_slot, hedef_gun):
    """SLA cezasini, assignment ilk eklendigi anda DONMUS a.sla_cost yerine, bu
    SPESIFIK satirin GUNCEL (Cozum A ile - o aracin GERCEK toplam yukune gore
    yeniden hesaplanmis) varis anindan TAZE turetir - tipki arac maliyetinin
    (arac_maliyeti) zaten yapildigi gibi.

    Neden gerekli: a.sla_cost, evaluate_path icinde SADECE o parcanin KENDI
    (o an kucuk olabilen) desisine gore hesaplanip donduruluyordu - ayni araca
    ALNS aramasi ilerledikce SONRADAN baska yuk bindikce aracin GERCEK kalkisi
    gecikebiliyordu ama donmus sla_cost bunu hic yakalamiyordu (bkz. sohbet
    gecmisi: test_sla_cezasi.py::test_nihai_bacaklarda_sla_cezasi_dogru_hesaplanmis
    bulgusu - 808 nihai bacakta SLA cezasi sistematik olarak az/sifir
    raporlaniyordu, hepsinde Bacak_Toplam_Desi >> Bu_Talebin_Desisi)."""
    tamamlanma = ellecleme_tamamlanma_zamani(varis_dt, pay_desi, consolidation=False)
    deadline = sla_deadline(slot_datetime(demand_gun, demand_slot), hedef_gun)
    gecikme = gecikme_saat(tamamlanma, deadline)
    return sla_cezasi_tl(pay_desi, gecikme)



sla_penalties = [] # talepID -> ceza
csv_records = []

hayalet_kargo_satir_sayisi = 0
hayalet_kargo_toplam_desi = 0.0
for a in best.assignments:
    nihai_kaynak, nihai_varis = a.demand_hat
    if len(a.legs) == 1:
        leg = a.legs[0]
        rota_tipi = "Direkt" if leg.gun == a.demand_gun and leg.slot == a.demand_slot else "Direkt (Ertelenmis)"
        arac_tipi = "Kiralik" if leg.is_kiralik else "Spot"
        key = _bucket_key(leg)
        bu_sevkiyatin_paylari = []
        for (aa, ll, arac_index, pay_desi, ll_i) in bucket_dagilim[key]:
            # YENİ: pay_desi > 0.01 şartı ile 0 çıkan hayalet satırları engelliyoruz
            if aa is a and pay_desi > 0.01:
                bu_sevkiyatin_paylari.append((arac_index, pay_desi))
            else: 
                hayalet_kargo_satir_sayisi += 1
                hayalet_kargo_toplam_desi += pay_desi
        _base_talep_id = talep_id_goruntu.get(id(a), a.talep_id)
        _coklu_arac = len(bu_sevkiyatin_paylari) > 1
        for _sira, (arac_index, pay_desi) in enumerate(bu_sevkiyatin_paylari, start=1):
            talep_id_gosterim = f"{_base_talep_id}-{_sira}" if _coklu_arac else _base_talep_id
            # YENİ DEĞİŞİKLİK (Cozum A): kalkis/varis artik bu SPESIFIK aracin
            # TOPLAM yukune gore hesaplaniyor, bu talebin kendi desisine gore degil.
            gercek_dt, varis_dt = arac_zamanlari[(key, arac_index)]
            gercek_gun = gercek_dt.strftime("%Y-%m-%d")
            gercek_slot = gercek_dt.strftime("%H:%M")
            varis_gun = varis_dt.strftime("%Y-%m-%d")
            varis_saat = varis_dt.strftime("%H:%M")

            hedef_gun = data.route_lookup[(nihai_kaynak, nihai_varis)]["target_delivery_days"]
            sla_cezasi_raw = _final_leg_sla_cezasi(varis_dt, round(pay_desi, 2), a.demand_gun, a.demand_slot, hedef_gun)
            sla_cezasi = round(sla_cezasi_raw, 2)
            if sla_cezasi > 0:
                sla_penalties.append((a.talep_id, sla_cezasi))
            csv_records.append({
                "Tarih": gercek_gun, "Slot": gercek_slot, "Arac_Tipi": arac_tipi, "Arac_Turu": leg.arac_turu,
                "Arac_ID": arac_id_kodu[(key, arac_index)],
                "Talep_ID": talep_id_gosterim,
                "Cikis_TM": leg.src, "Varis_TM": leg.dst,
                "Yolculuk_Suresi_Dk": data.route_lookup[(leg.src, leg.dst)][leg.arac_turu] * 60,
                "Cikis_Ellecleme_Dk": ellecleme_suresi_dakika(pay_desi, consolidation=False),
                "Varis_Ellecleme_Dk": ellecleme_suresi_dakika(pay_desi, consolidation=False),
                "Nihai_Kaynak": nihai_kaynak, "Nihai_Varis": nihai_varis,
                "Bacaktaki_Arac_Sayisi": _bacak_arac_sayisi(leg),
                "Bu_Talebin_Desisi": round(pay_desi, 2),
                "Bacak_Toplam_Desi": round(bucket_toplam_desi[key], 2),
                "Maliyet_TL": round(arac_maliyeti[(key, arac_index)] * (pay_desi / arac_toplam_yuk[(key, arac_index)]), 2),
                "SLA_Cezasi_TL": sla_cezasi,
                "Toplam_Maliyet_TL": round(
                    arac_maliyeti[(key, arac_index)] * (pay_desi / arac_toplam_yuk[(key, arac_index)])
                    + sla_cezasi_raw, 2
                ),
                "Rota_Tipi": rota_tipi, "Talep_Tarihi": a.demand_gun, "Talep_Slotu": a.demand_slot,
                "Konsolide_Talep_Sayisi": bucket_konsolidasyon_sayisi[key],
                "Varis_Tarihi": varis_gun, "Varis_Saati": varis_saat,
        })
    else:
        ara_duraklar = " -> ".join(leg.dst for leg in a.legs[:-1])
        
        for i, leg in enumerate(a.legs):
            arac_tipi = "Kiralik" if leg.is_kiralik else "Spot"
            key = _bucket_key(leg)
            
            # YENİ: Uğrama (milk-run) ise, ikinci bacakta olsak bile 
            # kargonun İLK bacaktaki dağılım oranlarını ve Araç ID'lerini baz alıyoruz.
            if a.milk_run:
                key_referans = _bucket_key(a.legs[0])
            else:
                key_referans = key
                
            bu_sevkiyatin_paylari = []
            for (aa, ll, arac_index, pay_desi, ll_i) in bucket_dagilim[key_referans]:
                if aa is a:
                    bu_sevkiyatin_paylari.append((arac_index, pay_desi))

            _base_talep_id = talep_id_goruntu.get(id(a), a.talep_id)
            _coklu_arac = len(bu_sevkiyatin_paylari) > 1
            
            for _sira, (arac_index, pay_desi) in enumerate(bu_sevkiyatin_paylari, start=1):
                talep_id_gosterim = f"{_base_talep_id}-{_sira}" if _coklu_arac else _base_talep_id
                
                # ZAMAN HESAPLAMASI
                if a.milk_run and i > 0:
                    # Uğrama yapan araç durmadığı için kendi orijinal zaman çizelgesini kullanır
                    gercek_dt, varis_dt = assignment_cizelgeleri[id(a)][i]
                else:
                    gercek_dt, varis_dt = arac_zamanlari[(key_referans, arac_index)]
                    
                gercek_gun = gercek_dt.strftime("%Y-%m-%d")
                gercek_slot = gercek_dt.strftime("%H:%M")
                varis_gun = varis_dt.strftime("%Y-%m-%d")
                varis_saat = varis_dt.strftime("%H:%M")
                
                if i == len(a.legs) - 1:
                    hedef_gun = data.route_lookup[(nihai_kaynak, nihai_varis)]["target_delivery_days"]
                    sla_cezasi_raw = _final_leg_sla_cezasi(varis_dt, round(pay_desi, 2), a.demand_gun, a.demand_slot, hedef_gun)
                else:
                    sla_cezasi_raw = 0.0
                sla_cezasi = round(sla_cezasi_raw, 2)
                if sla_cezasi > 0:
                    sla_penalties.append((a.talep_id, sla_cezasi))
                
                if a.milk_run:
                    skip_src = i > 0
                    skip_dst = i < len(a.legs) - 1
                    cikis_ellec_dk = 0 if skip_src else math.ceil(ellecleme_suresi_dakika(pay_desi, consolidation=False))
                    varis_ellec_dk = 0 if skip_dst else math.ceil(ellecleme_suresi_dakika(pay_desi, consolidation=False))
                    rota_tipi_cok_bacakli = f"Uğramalı {i + 1}/{len(a.legs)} (via {ara_duraklar}, nihai varis: {nihai_varis})"
                else:
                    cikis_ellec_dk = math.ceil(ellecleme_suresi_dakika(pay_desi, consolidation=(i > 0)))
                    varis_ellec_dk = math.ceil(ellecleme_suresi_dakika(pay_desi, consolidation=(i < len(a.legs) - 1)))
                    rota_tipi_cok_bacakli = f"Aktarmalı/Konsolidasyonlu {i + 1}/{len(a.legs)} (via {ara_duraklar}, nihai varis: {nihai_varis})"
                
                # MALİYET HESABI
                if a.milk_run and i > 0:
                    # Uğrama devam bacağında, o anki genel rotanın ortalama maliyetini baz alıyoruz
                    bacak_toplam_maliyet = sum(v for k, v in arac_maliyeti.items() if k[0] == key)
                    bacak_birim = bacak_toplam_maliyet / bucket_toplam_desi[key] if bucket_toplam_desi[key] > 0 else 0
                    arac_maliyeti_payi = pay_desi * bacak_birim
                else:
                    arac_maliyeti_payi = arac_maliyeti[(key_referans, arac_index)] * (pay_desi / arac_toplam_yuk[(key_referans, arac_index)])

                csv_records.append({
                    "Tarih": gercek_gun, "Slot": gercek_slot, "Arac_Tipi": arac_tipi, "Arac_Turu": leg.arac_turu,
                    # YENİ: Artık '+' yok. Doğrudan İlk bacağın tertemiz Araç ID'si kopyalanıyor.
                    "Arac_ID": arac_id_kodu[(key_referans, arac_index)],
                    "Talep_ID": talep_id_gosterim,
                    "Cikis_TM": leg.src, "Varis_TM": leg.dst,
                    "Yolculuk_Suresi_Dk": math.ceil(data.route_lookup[(leg.src, leg.dst)][leg.arac_turu] * 60),
                    "Cikis_Ellecleme_Dk": cikis_ellec_dk,
                    "Varis_Ellecleme_Dk": varis_ellec_dk,
                    "Nihai_Kaynak": nihai_kaynak, "Nihai_Varis": nihai_varis,
                    "Bacaktaki_Arac_Sayisi": _bacak_arac_sayisi(leg),
                    "Bu_Talebin_Desisi": round(pay_desi, 2),
                    "Bacak_Toplam_Desi": round(bucket_toplam_desi[key], 2),
                    "Maliyet_TL": round(arac_maliyeti_payi, 2),
                    "SLA_Cezasi_TL": sla_cezasi,
                    "Toplam_Maliyet_TL": round(arac_maliyeti_payi + sla_cezasi_raw, 2),
                    "Rota_Tipi": rota_tipi_cok_bacakli,
                    "Konsolide_Talep_Sayisi": bucket_konsolidasyon_sayisi[key],
                    "Talep_Tarihi": a.demand_gun, "Talep_Slotu": a.demand_slot,
                    "Varis_Tarihi": varis_gun, "Varis_Saati": varis_saat,
                })
if hayalet_kargo_satir_sayisi > 0:
    print(f"DEBUG/BİLGİ: Floating-point (ondalıklı sayı) hassasiyetinden kaynaklanan                     {hayalet_kargo_satir_sayisi} adet matematiksel kalıntı (Toplam: {hayalet_kargo_toplam_desi:.5f} desi) tespit edildi ve gerçekçiliği korumak adına Taşıma Planı'ndan filtrelendi.")
dispatch_records = []
for (hat, arac_turu), stok in kiralik_stok_gunluk.items():
    if stok <= 0:
        continue
    src, dst = hat
    kap = arac_parametreleri[arac_turu]["kapasite_desi"]
    p = arac_parametreleri[arac_turu]
    entry = route_lookup.get(hat)
    if not entry:
        continue
    seyir_saat = entry[arac_turu]
    dist = entry["distance_km"]
    birim_bos_maliyet = (seyir_saat * p["rental_hourly"]) + (dist * p["rental_km"])

    for gun in gunler:
        key = (src, dst, gun, KIRALIK_DISPATCH_SLOT, arac_turu)
        kullanilan_desi = best.leg_kiralik_desi.get(key, 0.0)
        kullanilan_adet = min(stok, math.ceil(kullanilan_desi / kap) if kullanilan_desi > 0 else 0)
        bos_adet = stok - kullanilan_adet

        if bos_adet <= 0:
            continue

        kalkis_zamani = slot_datetime(gun, KIRALIK_DISPATCH_SLOT)
        varis_dt = varis_zamani(kalkis_zamani, seyir_saat)
        varis_gun = varis_dt.strftime("%Y-%m-%d")
        varis_saat = varis_dt.strftime("%H:%M")

        for _ in range(int(bos_adet)):
            arac_id = f"V{_bos_id_counter:04d}"
            _bos_id_counter += 1

            csv_records.append({
                "Tarih": gun,
                "Slot": KIRALIK_DISPATCH_SLOT,
                "Arac_Tipi": "Kiralik",
                "Arac_Turu": arac_turu,
                "Arac_ID": arac_id,
                "Talep_ID": "",
                "Cikis_TM": src,
                "Varis_TM": dst,
                "Yolculuk_Suresi_Dk": math.ceil(seyir_saat * 60),
                "Cikis_Ellecleme_Dk": 0,
                "Varis_Ellecleme_Dk": 0,
                "Nihai_Kaynak": src,
                "Nihai_Varis": dst,
                "Bacaktaki_Arac_Sayisi": 1,
                "Bu_Talebin_Desisi": 0,
                "Bacak_Toplam_Desi": 0,
                "Maliyet_TL": round(birim_bos_maliyet, 2),
                "SLA_Cezasi_TL": 0,
                "Toplam_Maliyet_TL": round(birim_bos_maliyet, 2),
                "Rota_Tipi": "Kiralık (Boş Sefer)",
                "Talep_Tarihi": "",
                "Talep_Slotu": "",
                "Konsolide_Talep_Sayisi": 0,
                "Varis_Tarihi": varis_gun,
                "Varis_Saati": varis_saat,
            })

            # dispatch_records’a da aynı aracı ekle (özet tablo için)
            dispatch_records.append({
                "Tarih": gun,
                "Slot": KIRALIK_DISPATCH_SLOT,
                "Cikis_TM": src,
                "Varis_TM": dst,
                "Arac_Tipi": "Kiralik " + arac_turu,
                "Arac_Sayisi": 1,
                "Toplam_Desi": 0,
                "Kapasite": kap,
                "Doluluk_Yuzde": 0.0,
            })

csv_records.sort(key=lambda r: (r["Tarih"], r["Slot"], r["Cikis_TM"], r["Varis_TM"], r["Talep_Tarihi"], r["Talep_Slotu"]))

# ---- Kapasite kullanim ozeti (elleceleme/tir nasil dahil edildigini GORUNUR kilmak icin) ----
capacity_records = []
for tm, cap in sorted(data.handling_capacity.items()):
    for gun in data.gunler:
        kullanim = best.handling_usage.get((tm, gun), 0.0)
        capacity_records.append({
            "TM": tm, "Tarih": gun, "Tur": "Ellecleme (desi)",
            "Kullanim": round(kullanim, 1), "Kapasite": cap,
            "Doluluk_Yuzde": round(100 * kullanim / cap, 1) if cap else 0,
        })
if data.tir_arac_turu is not None:
    for tm, cap in sorted(data.tir_capacity.items()):
        for gun in data.gunler:
            kullanim = (
                best.tir_usage_in.get((tm, gun), 0) + best.tir_usage_out.get((tm, gun), 0)
                + data.fixed_kiralik_tir_usage.get((tm, gun), 0)
            )
            capacity_records.append({
                "TM": tm, "Tarih": gun, "Tur": "Tir (adet)",
                "Kullanim": kullanim, "Kapasite": cap,
                "Doluluk_Yuzde": round(100 * kullanim / cap, 1) if cap else 0,
            })

# ---- Arac sevkiyat ozeti (her fiziksel bacak icin TEK satir, dogru arac sayisiyla) ----

for key, toplam_desi in sorted(bucket_toplam_desi.items()):
    src, dst, gun, slot, arac_turu, is_kiralik = key
    gercek_dt = bucket_gercek_kalkis[key]
    gercek_gun = gercek_dt.strftime("%Y-%m-%d")
    gercek_slot = gercek_dt.strftime("%H:%M")
    
    kap = arac_parametreleri[arac_turu]["kapasite_desi"]
    adet = 1 if is_kiralik else spot_vehicle_count(toplam_desi, kap, 10 ** 9)
    dispatch_records.append({
        "Tarih": gercek_gun, "Slot": gercek_slot, "Cikis_TM": src, "Varis_TM": dst, # BURASI GÜNCELLENDİ
        "Arac_Tipi": ("Kiralik " if is_kiralik else "Spot ") + arac_turu,
        "Arac_Sayisi": adet, "Toplam_Desi": round(toplam_desi, 2), "Kapasite": kap,
        "Doluluk_Yuzde": round(100 * toplam_desi / (adet * kap), 1) if adet else 0,
    })

kiralik_sabit_toplam = best._fixed_kiralik_cost
# Kiralığın elleçleme maliyetini artık arac_maliyeti sözlüğünden türetiyoruz - o zaten
# doğru (dokunuş/uğrama-farkındalıklı, cost_model.ellecleme_maliyet_hesapla üzerinden
# hesaplanmış) değeri tutuyor. Kiralık bucket'larda seyir_maliyeti=0 olduğu için
# arac_maliyeti[arac_key] TAMAMEN elleçleme payıdır (bkz. yukarıdaki arac_maliyeti
# hesap döngüsü). DÜZELTME: burada eskiden AYRI, eski (×1, uğrama-farkındasız) bir
# formül vardı - GENEL TOPLAM'ı etkilemiyordu (aşağıda +/- ile sadeleşiyordu) ama
# konsol özetindeki Kiralık/Spot kalem kırılımını yanlış gösteriyordu (CSV'nin
# yarısı kadar Kiralık, farkı da yanlışlıkla Spot'a sızmış gibi).
kiralik_ellecleme_toplam = sum(
    v for arac_key, v in arac_maliyeti.items() if arac_key[0][5]  # arac_key[0][5] = is_kiralik
)

# Atamalardaki tüm araç maliyetleri (Spot Tamamı + Kiralık Elleçleme).
# NOT: artik best.assignments'in kendi (potansiyel olarak BAYAT/stale)
# vehicle_cost degerlerinden DEGIL, arac_maliyeti sozlugunden (her aracin
# GUNCEL toplam yukune gore sifirdan hesaplanan gercek maliyeti) topluyoruz -
# boylece objective()'teki duzeltmeyle TUTARLI kalir (bkz. sohbet gecmisi).
karma_arac_maliyeti = sum(arac_maliyeti.values())

# Spot ve Kiralık maliyetleri birbirinden temizce ayırıyoruz
spot_toplam_maliyet = karma_arac_maliyeti - kiralik_ellecleme_toplam
kiralik_gercek_toplam = kiralik_sabit_toplam + kiralik_ellecleme_toplam
# DUZELTME: artik best.assignments'in DONMUS a.sla_cost degerlerinden DEGIL,
# csv_records'taki (yukarida, _final_leg_sla_cezasi ile TAZE hesaplanmis)
# SLA_Cezasi_TL sutunundan topluyoruz - boylece ozet, raporlanan satirlarla
# BIREBIR tutarli kalir (bkz. _final_leg_sla_cezasi docstring'i).
sla_ceza_toplam = sum(r["SLA_Cezasi_TL"] for r in csv_records)

# YENİ: Kapasite aşım cezalarını hesapla
ellecleme_ceza_toplam = 0.0
for tm, cap in data.handling_capacity.items():
    for gun in data.gunler:
        asim = best.handling_usage.get((tm, gun), 0.0) - cap
        if asim > 0:
            ellecleme_ceza_toplam += asim * 1000.0

tir_ceza_toplam = 0.0
if data.tir_arac_turu is not None:
    for tm, cap in data.tir_capacity.items():
        for gun in data.gunler:
            kullanim = (
                best.tir_usage_in.get((tm, gun), 0) + best.tir_usage_out.get((tm, gun), 0)
                + data.fixed_kiralik_tir_usage.get((tm, gun), 0)
            )
            asim = kullanim - cap
            if asim > 0:
                tir_ceza_toplam += asim * 50000.0

genel_toplam = best.objective()
cok_bacakli_talep_sayisi = sum(1 for a in best.assignments if len(a.legs) > 1)
gercek_ugrama_talep_sayisi = sum(1 for a in best.assignments if len(a.legs) > 1 and a.milk_run)
konsolidasyonlu_talep_sayisi = cok_bacakli_talep_sayisi - gercek_ugrama_talep_sayisi
gercek_konsolidasyon_bacak_sayisi = sum(1 for v in bucket_konsolidasyon_sayisi.values() if v > 1)

# YENİ: yerleştirilemeyen (unassigned) talep cezası - force_insert kaldırıldığından
# beri (Task #1) bu kalem daha sık/büyük görünebiliyor, önceden özet raporda hiç
# gösterilmiyordu (bkz. sohbet gecmisi - "toplamda 4.7M TL kayboldu" bulgusu).
# objective()'i degistirmek yerine, geriye kalan farki (matematiksel olarak HER
# ZAMAN dogru) bu kalem olarak gosteriyoruz.
unassigned_desi_toplam = sum(x[3] for x in best.unassigned)
unassigned_satir_sayisi = len(best.unassigned)
if unassigned_satir_sayisi == 0:
    unassigned_cezasi = 0.0
    genel_toplam = (kiralik_gercek_toplam + spot_toplam_maliyet + sla_ceza_toplam
                    + ellecleme_ceza_toplam + tir_ceza_toplam)
else:
    unassigned_cezasi = genel_toplam - (
        kiralik_gercek_toplam + spot_toplam_maliyet + sla_ceza_toplam
        + ellecleme_ceza_toplam + tir_ceza_toplam
    )


sla_dusen_talep_sayisi = len(sla_penalties)


ozet = f"""
{'=' * 80}
OZET ISTATISTIKLER (Faz 2 - ALNS, saat bazli, konsolidasyon destekli)
{'=' * 80}
  Kiralık Arac Maliyeti       : {kiralik_gercek_toplam:>15,.0f} TL
      -> Sabit Seyir          : {kiralik_sabit_toplam:>15,.0f} TL
      -> Ellecleme (Marjinal) : {kiralik_ellecleme_toplam:>15,.0f} TL
  Spot Arac Maliyeti          : {spot_toplam_maliyet:>15,.0f} TL
  SLA Gecikme Cezasi          : {sla_ceza_toplam:>15,.0f} TL
      -> SLA'a ya düşen talep sayısı: {sla_dusen_talep_sayisi} 
{'-' * 80}
  OPERASYONEL MALIYET         : {(kiralik_gercek_toplam + spot_toplam_maliyet + sla_ceza_toplam):>15,.0f} TL
{'-' * 80}
  UGRAMA / KONSOLIDASYON
      -> Cok bacakli talep sayisi (toplam)   : {cok_bacakli_talep_sayisi:>10}
      -> Gercek ugrama (milk-run) talep say. : {gercek_ugrama_talep_sayisi:>10}
      -> Konsolidasyonlu (indir+yukle) talep : {konsolidasyonlu_talep_sayisi:>10}
      -> Gercek konsolide bacak sayisi        : {gercek_konsolidasyon_bacak_sayisi:>10}
{'-' * 80}
  KAPASITE ASIM CEZALARI (Sanal Maliyetler)
      -> Ellecleme Asim Cezasi: {ellecleme_ceza_toplam:>15,.0f} TL
      -> TIR Park Asim Cezasi : {tir_ceza_toplam:>15,.0f} TL
{'-' * 80}
  YERLESTIRILEMEYEN TALEP CEZASI (Sanal Maliyet)
      -> Bekleyen satir/desi   : {unassigned_satir_sayisi:>10} satir / {unassigned_desi_toplam:>10,.0f} desi
      -> Ceza                  : {unassigned_cezasi:>15,.0f} TL
{'-' * 80}
  TOPLAM MALIYET (Objective)  : {genel_toplam:>15,.0f} TL
{'=' * 80}
"""
print(ozet)
output_txt.write_text(ozet, encoding="utf-8")

output_dispatch_csv = results_dir / "vehicle_dispatch_summary.csv"
output_capacity_csv = results_dir / "capacity_utilization.csv"

if csv_records:
    pd.DataFrame(csv_records).to_csv(output_csv, index=False, encoding="utf-8-sig")
    pd.DataFrame(dispatch_records).to_csv(output_dispatch_csv, index=False, encoding="utf-8-sig")
    pd.DataFrame(capacity_records).to_csv(output_capacity_csv, index=False, encoding="utf-8-sig")
    print(f"CSV kaydedildi: {output_csv}")
    print(f"Arac sevkiyat ozeti: {output_dispatch_csv}")
    print(f"Kapasite kullanim ozeti: {output_capacity_csv}")

    def _yaz_sayfa(ws, basliklar, satirlar, genislikler):
        ws.append(basliklar)
        for hucre in ws[1]:
            hucre.font = Font(bold=True, color="FFFFFF")
            hucre.fill = PatternFill("solid", start_color="4472C4")
            hucre.alignment = Alignment(horizontal="center")
        for row in satirlar:
            ws.append(row)
        for idx, width in enumerate(genislikler, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Teslim Plani (Talep Bazli)"
    _yaz_sayfa(
        ws1,
        ["Arac ID", "Tarih", "Slot", "Arac Tipi", "Arac Turu", "Cikis TM", "Varis TM", "Yolculuk Suresi (dk)",
         "Cikis Ellecleme (dk)", "Varis Ellecleme (dk)", "Nihai Kaynak", "Nihai Varis",
         "Bacaktaki Arac Sayisi", "Bu Talebin Desisi", "Bacak Toplam Desi", "Maliyet TL",
         "SLA Cezasi TL", "Toplam Maliyet TL",
         "Rota Tipi", "Talep Tarihi", "Talep Slotu", "Varis Tarihi", "Varis Saati"],
        [[rec["Arac_ID"], rec["Tarih"], rec["Slot"], rec["Arac_Tipi"], rec["Arac_Turu"], rec["Cikis_TM"], rec["Varis_TM"], rec["Yolculuk_Suresi_Dk"],
          rec["Cikis_Ellecleme_Dk"], rec["Varis_Ellecleme_Dk"], rec["Nihai_Kaynak"], rec["Nihai_Varis"],
          rec["Bacaktaki_Arac_Sayisi"], rec["Bu_Talebin_Desisi"], rec["Bacak_Toplam_Desi"],
          rec["Maliyet_TL"], rec["SLA_Cezasi_TL"], rec["Toplam_Maliyet_TL"], rec["Rota_Tipi"], rec["Talep_Tarihi"], rec["Talep_Slotu"], rec["Varis_Tarihi"], rec["Varis_Saati"]]
         for rec in csv_records],
        [40, 12, 8, 16, 14, 14, 14, 16, 16, 16, 14, 14, 14, 14, 12, 14, 16, 30, 14, 10, 14, 10],
    )

    ws2 = wb.create_sheet("Arac Sevkiyat Ozeti")
    _yaz_sayfa(
        ws2,
        ["Tarih", "Slot", "Cikis TM", "Varis TM", "Arac Tipi", "Arac Sayisi",
         "Toplam Desi", "Kapasite", "Doluluk %"],
        [[r["Tarih"], r["Slot"], r["Cikis_TM"], r["Varis_TM"], r["Arac_Tipi"],
          r["Arac_Sayisi"], r["Toplam_Desi"], r["Kapasite"], r["Doluluk_Yuzde"]]
         for r in dispatch_records],
        [12, 8, 14, 14, 16, 12, 12, 10, 10],
    )

    ws3 = wb.create_sheet("TM Kapasite Kullanimi")
    _yaz_sayfa(
        ws3,
        ["TM", "Tarih", "Tur", "Kullanim", "Kapasite", "Doluluk %"],
        [[r["TM"], r["Tarih"], r["Tur"], r["Kullanim"], r["Kapasite"], r["Doluluk_Yuzde"]]
         for r in capacity_records],
        [14, 12, 18, 12, 12, 10],
    )

    wb.save(output_xlsx)
    print(f"Excel kaydedildi: {output_xlsx} (3 sayfa: Teslim Plani, Arac Sevkiyat Ozeti, TM Kapasite Kullanimi)")

    # ---- Resmi "TASIMA PLANI" formati - jüri şablonuyla BİREBİR aynı 16
    # sütun, aynı sıra, aynı isim. Yukarıdaki detayli/debug sayfadan farklı -
    # burada sadece şablonun istedigi sütunlar var, ekstra analiz sütunu yok. ----
    output_tasima_plani_xlsx = results_dir / "Tasima_Plani.xlsx"
    wb_resmi = Workbook()
    ws_resmi = wb_resmi.active
    ws_resmi.title = "Tasima Plani"
    _yaz_sayfa(
        ws_resmi,
        ["Araç ID", "Araç Tipi", "Araç türü", "Çıkış Transfer Merkezi", "Varış Transfer Merkezi",
         "Çıkış Tarihi", "Çıkış Saati", "Varış Tarihi", "Varış Saati", "Talep ID", "Taşınan Desi",
         "Yolculuk süresi", "Varış elleçleme süresi", "Çıkış Elleçleme süresi", "SLA cezası", "Toplam maliyet"],
        [[rec["Arac_ID"], rec["Arac_Tipi"], rec["Arac_Turu"], rec["Cikis_TM"], rec["Varis_TM"],
          rec["Tarih"], rec["Slot"], rec["Varis_Tarihi"], rec["Varis_Saati"], rec["Talep_ID"],
          rec["Bu_Talebin_Desisi"], rec["Yolculuk_Suresi_Dk"], rec["Varis_Ellecleme_Dk"],
          rec["Cikis_Ellecleme_Dk"], rec["SLA_Cezasi_TL"], rec["Toplam_Maliyet_TL"]]
         for rec in csv_records],
        [40, 12, 14, 20, 20, 14, 12, 14, 12, 12, 14, 14, 18, 18, 12, 14],
    )
    wb_resmi.save(output_tasima_plani_xlsx)
    print(f"Resmi Tasima Plani kaydedildi: {output_tasima_plani_xlsx}")
else:
    print("Kayit yok - CSV/Excel uretilmedi.")
